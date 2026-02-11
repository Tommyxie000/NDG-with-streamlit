import streamlit as st
import numpy as np
import matplotlib.pyplot as plt
import pandas as pd
from datetime import datetime
import io
import base64

# 提前导入Excel相关库
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.drawing.image import Image
from openpyxl.utils import get_column_letter
from io import BytesIO

# 设置matplotlib字体，解决中文显示问题
plt.rcParams['font.sans-serif'] = ['Microsoft YaHei', 'Arial Unicode MS', 'SimHei', 'DejaVu Sans']
plt.rcParams['axes.unicode_minus'] = False

# 导入统计模块
from statistics import generate_numbers_with_cpk, generate_numbers_robust_spc, calculate_additional_stats, perform_normality_tests, InputValidationError, GenerationError, StatisticsError

# 设置页面配置
st.set_page_config(
    page_title="NDG - 正态分布随机数生成器",
    page_icon="📊",
    layout="wide",
    initial_sidebar_state="expanded"
)

# 自定义CSS样式 - 专业质量管理风格
st.markdown("""
<style>
    /* 全局样式 */
    .main {
        padding: 0rem 1rem;
    }
    
    /* 头部样式 */
    .app-header {
        background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
        padding: 1.5rem 2rem;
        border-radius: 12px;
        margin-bottom: 1.5rem;
        box-shadow: 0 4px 15px rgba(0,0,0,0.1);
    }
    .app-title {
        font-size: 2rem;
        font-weight: 700;
        color: white;
        margin: 0;
        text-shadow: 1px 1px 2px rgba(0,0,0,0.2);
    }
    .app-subtitle {
        font-size: 0.95rem;
        color: rgba(255,255,255,0.9);
        margin-top: 0.3rem;
    }
    
    /* 控制面板样式 */
    .control-panel {
        background: white;
        padding: 1.5rem;
        border-radius: 12px;
        box-shadow: 0 2px 8px rgba(0,0,0,0.08);
        margin-bottom: 1.5rem;
        border: 1px solid #e8e8e8;
    }
    .panel-title {
        font-size: 1.1rem;
        font-weight: 600;
        color: #2c3e50;
        margin-bottom: 1rem;
        padding-bottom: 0.5rem;
        border-bottom: 2px solid #667eea;
    }
    
    /* 条件组标签样式 */
    .group-tabs {
        display: flex;
        gap: 0.5rem;
        margin-bottom: 1rem;
        flex-wrap: wrap;
    }
    .group-tab {
        padding: 0.5rem 1rem;
        background: #f0f2f5;
        border-radius: 20px;
        font-size: 0.9rem;
        color: #555;
        cursor: pointer;
        transition: all 0.3s;
        border: 2px solid transparent;
    }
    .group-tab:hover {
        background: #e8eaf6;
    }
    .group-tab.active {
        background: #667eea;
        color: white;
        border-color: #5a6fd6;
    }
    .group-tab.has-data {
        border-color: #27ae60;
    }
    
    /* 统计卡片样式 */
    .stat-grid {
        display: grid;
        grid-template-columns: repeat(auto-fit, minmax(120px, 1fr));
        gap: 0.8rem;
        margin-bottom: 1.5rem;
    }
    .stat-card {
        background: linear-gradient(135deg, #f5f7fa 0%, #e4e8ec 100%);
        padding: 1rem;
        border-radius: 10px;
        text-align: center;
        border-left: 4px solid #667eea;
        transition: transform 0.2s;
    }
    .stat-card:hover {
        transform: translateY(-2px);
        box-shadow: 0 4px 12px rgba(0,0,0,0.1);
    }
    .stat-card.cpk-good {
        border-left-color: #27ae60;
    }
    .stat-card.cpk-warning {
        border-left-color: #f39c12;
    }
    .stat-card.cpk-danger {
        border-left-color: #e74c3c;
    }
    .stat-label {
        font-size: 0.75rem;
        color: #7f8c8d;
        text-transform: uppercase;
        letter-spacing: 0.5px;
        margin-bottom: 0.3rem;
    }
    .stat-value {
        font-size: 1.4rem;
        font-weight: 700;
        color: #2c3e50;
    }
    .stat-unit {
        font-size: 0.75rem;
        color: #95a5a6;
    }
    
    /* 对比表格样式 */
    .comparison-table {
        width: 100%;
        border-collapse: collapse;
        margin-top: 1rem;
        font-size: 0.9rem;
    }
    .comparison-table th {
        background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
        color: white;
        padding: 0.8rem;
        text-align: center;
        font-weight: 600;
        position: sticky;
        top: 0;
    }
    .comparison-table td {
        padding: 0.7rem;
        text-align: center;
        border-bottom: 1px solid #ecf0f1;
    }
    .comparison-table tr:nth-child(even) {
        background: #f8f9fa;
    }
    .comparison-table tr:hover {
        background: #e8eaf6;
    }
    .comparison-table .group-name {
        font-weight: 600;
        color: #2c3e50;
    }
    .comparison-table .metric-good {
        color: #27ae60;
        font-weight: 600;
    }
    .comparison-table .metric-warning {
        color: #f39c12;
        font-weight: 600;
    }
    .comparison-table .metric-danger {
        color: #e74c3c;
        font-weight: 600;
    }
    
    /* 图表容器样式 */
    .chart-container {
        background: white;
        padding: 1.5rem;
        border-radius: 12px;
        box-shadow: 0 2px 8px rgba(0,0,0,0.08);
        margin-bottom: 1.5rem;
    }
    .chart-title {
        font-size: 1.1rem;
        font-weight: 600;
        color: #2c3e50;
        margin-bottom: 1rem;
        display: flex;
        align-items: center;
        gap: 0.5rem;
    }
    .chart-title::before {
        content: "📈";
    }
    
    /* 数据表格容器 */
    .data-section {
        background: white;
        padding: 1.5rem;
        border-radius: 12px;
        box-shadow: 0 2px 8px rgba(0,0,0,0.08);
        margin-bottom: 1.5rem;
    }
    
    /* 操作按钮样式 */
    .action-bar {
        display: flex;
        gap: 0.8rem;
        flex-wrap: wrap;
        margin-bottom: 1.5rem;
    }
    
    /* 侧边栏样式优化 */
    .sidebar-content {
        padding: 1rem 0;
    }
    .sidebar-section {
        background: white;
        padding: 1rem;
        border-radius: 10px;
        margin-bottom: 1rem;
        border: 1px solid #e8e8e8;
    }
    .sidebar-section-title {
        font-size: 0.95rem;
        font-weight: 600;
        color: #2c3e50;
        margin-bottom: 0.8rem;
        padding-bottom: 0.5rem;
        border-bottom: 2px solid #667eea;
    }
    
    /* 空状态样式 */
    .empty-state {
        text-align: center;
        padding: 3rem;
        color: #95a5a6;
    }
    .empty-state-icon {
        font-size: 4rem;
        margin-bottom: 1rem;
    }
    
    /* 滚动容器 */
    .scroll-container {
        overflow-x: auto;
        padding-bottom: 0.5rem;
    }
    .scroll-container::-webkit-scrollbar {
        height: 8px;
    }
    .scroll-container::-webkit-scrollbar-track {
        background: #f1f1f1;
        border-radius: 4px;
    }
    .scroll-container::-webkit-scrollbar-thumb {
        background: #c1c1c1;
        border-radius: 4px;
    }
    .scroll-container::-webkit-scrollbar-thumb:hover {
        background: #a8a8a8;
    }
    
    /* 质量指标徽章 */
    .quality-badge {
        display: inline-block;
        padding: 0.25rem 0.75rem;
        border-radius: 12px;
        font-size: 0.8rem;
        font-weight: 600;
    }
    .quality-badge.excellent {
        background: #d4edda;
        color: #155724;
    }
    .quality-badge.good {
        background: #fff3cd;
        color: #856404;
    }
    .quality-badge.poor {
        background: #f8d7da;
        color: #721c24;
    }
    
    /* 隐藏Streamlit默认元素 */
    #MainMenu {visibility: hidden;}
    footer {visibility: hidden;}
    .stDeployButton {display: none;}
</style>
""", unsafe_allow_html=True)

# 初始化会话状态
# 确保在任何运行环境下都能正确初始化
try:
    # 检查是否在Streamlit环境中运行
    if not hasattr(st, 'session_state'):
        # 非Streamlit环境，创建模拟会话状态
        class MockSessionState:
            def __init__(self):
                self.condition_groups = []
                self.current_group_index = -1
                self.selected_groups = []  # 存储用户选中的多个条件组索引
                self.show_comparison = False
                self.view_mode = "single"
        
        st.session_state = MockSessionState()
    else:
        # Streamlit环境，正常初始化
        if 'condition_groups' not in st.session_state:
            st.session_state.condition_groups = []
        if 'current_group_index' not in st.session_state:
            st.session_state.current_group_index = -1
        if 'selected_groups' not in st.session_state:
            st.session_state.selected_groups = []  # 存储用户选中的多个条件组索引
        if 'show_comparison' not in st.session_state:
            st.session_state.show_comparison = False
        if 'view_mode' not in st.session_state:
            st.session_state.view_mode = "single"  # single 或 comparison
except Exception:
    # 发生任何错误时，创建模拟会话状态
    class MockSessionState:
        def __init__(self):
            self.condition_groups = []
            self.current_group_index = -1
            self.selected_groups = []  # 存储用户选中的多个条件组索引
            self.show_comparison = False
            self.view_mode = "single"
    
    st.session_state = MockSessionState()

# 辅助函数：获取CPK状态
def get_cpk_status(cpk):
    if cpk >= 1.67:
        return "excellent", "优秀"
    elif cpk >= 1.33:
        return "good", "良好"
    else:
        return "poor", "需改进"

# 辅助函数：获取CPK颜色
def get_cpk_color(cpk):
    if cpk >= 1.67:
        return "#27ae60"
    elif cpk >= 1.33:
        return "#f39c12"
    else:
        return "#e74c3c"

# ==================== 页面头部 ====================
st.markdown("""
<div class="app-header">
    <h1 class="app-title">📊 NDG - 正态分布随机数生成器</h1>
    <div class="app-subtitle">专业质量工程统计分析工具 | SPC Process Control</div>
</div>
""", unsafe_allow_html=True)

# ==================== 侧边栏：参数设置 ====================
with st.sidebar:
    st.markdown('<div class="sidebar-content">', unsafe_allow_html=True)
    
    # 检查是否有条件组
    if st.session_state.condition_groups and st.session_state.current_group_index >= 0:
        # 当前选中的条件组
        current_group = st.session_state.condition_groups[st.session_state.current_group_index]
        current_params = current_group['params']
        
        # 条件组编辑
        st.markdown('<div class="sidebar-section">', unsafe_allow_html=True)
        st.markdown(f'<div class="sidebar-section-title">✏️ 编辑条件组：{current_group["name"]}</div>', unsafe_allow_html=True)
        
        # 核心参数编辑
        new_center = st.number_input("目标值 (Target)", value=current_params['center'], step=0.1, format="%.2f")
        new_upper_spec = st.number_input("规格上限 (USL)", value=current_params['upper_spec'], step=0.1, format="%.2f")
        new_lower_spec = st.number_input("规格下限 (LSL)", value=current_params['lower_spec'], step=0.1, format="%.2f")
        new_cpk_upper = st.number_input("CPK上限", value=current_params['cpk_upper'], step=0.01, format="%.2f")
        new_cpk_lower = st.number_input("CPK下限", value=current_params['cpk_lower'], step=0.01, format="%.2f")
        new_mean = st.number_input("均值 (Mean)", value=current_params['mean'], step=0.1, format="%.2f")
        
        # 其他参数编辑
        new_precision = st.slider("精度 (小数位)", min_value=0, max_value=6, value=current_params['precision'])
        new_count = st.slider("样本数量", min_value=1, max_value=500, value=current_params['count'])
        
        # 条件组名称编辑
        new_group_name = st.text_input("条件组名称", value=current_group['name'])
        
        # 保存修改按钮
        if st.button("💾 保存修改", use_container_width=True, type="primary"):
            if new_upper_spec <= new_lower_spec:
                st.error("规格上限必须大于规格下限！")
            elif new_cpk_upper <= new_cpk_lower:
                st.error("CPK上限必须大于CPK下限！")
            else:
                # 更新条件组信息
                current_group['name'] = new_group_name
                current_group['params'].update({
                    'center': new_center,
                    'upper_spec': new_upper_spec,
                    'lower_spec': new_lower_spec,
                    'cpk_upper': new_cpk_upper,
                    'cpk_lower': new_cpk_lower,
                    'mean': new_mean,
                    'precision': new_precision,
                    'count': new_count
                })
                # 清除已生成的数据，因为参数已更改
                current_group['numbers'] = None
                current_group['stats'] = None
                st.success(f"✓ 已更新：{new_group_name}")
                st.rerun()
        
        st.markdown('</div>', unsafe_allow_html=True)
        
        # 条件组管理
        st.markdown('<div class="sidebar-section">', unsafe_allow_html=True)
        st.markdown('<div class="sidebar-section-title">🗑️ 条件组管理</div>', unsafe_allow_html=True)
        
        if st.button("❌ 删除当前条件组", use_container_width=True, type="secondary"):
            deleted_group_name = current_group['name']
            st.session_state.condition_groups.pop(st.session_state.current_group_index)
            if st.session_state.condition_groups:
                st.session_state.current_group_index = min(st.session_state.current_group_index, len(st.session_state.condition_groups) - 1)
            else:
                st.session_state.current_group_index = -1
            st.success(f"✓ 已删除：{deleted_group_name}")
            st.rerun()
        
        st.markdown('</div>', unsafe_allow_html=True)
        
        # 新增条件组
        st.markdown('<div class="sidebar-section">', unsafe_allow_html=True)
        st.markdown('<div class="sidebar-section-title">➕ 新增条件组</div>', unsafe_allow_html=True)
        
        new_group_name = st.text_input("新条件组名称", value=f"条件组 {len(st.session_state.condition_groups) + 1}")
        
        # 新增条件组的参数设置
        new_group_center = st.number_input("目标值 (Target)", value=10.0, step=0.1, format="%.2f", key="new_center")
        new_group_upper_spec = st.number_input("规格上限 (USL)", value=12.0, step=0.1, format="%.2f", key="new_upper_spec")
        new_group_lower_spec = st.number_input("规格下限 (LSL)", value=8.0, step=0.1, format="%.2f", key="new_lower_spec")
        new_group_cpk_upper = st.number_input("CPK上限", value=1.67, step=0.01, format="%.2f", key="new_cpk_upper")
        new_group_cpk_lower = st.number_input("CPK下限", value=1.33, step=0.01, format="%.2f", key="new_cpk_lower")
        new_group_mean = st.number_input("均值 (Mean)", value=10.0, step=0.1, format="%.2f", key="new_mean")
        new_group_precision = st.slider("精度 (小数位)", min_value=0, max_value=6, value=2, key="new_precision")
        new_group_count = st.slider("样本数量", min_value=1, max_value=500, value=32, key="new_count")
        
        if st.button("添加新条件组", use_container_width=True, type="primary"):
            if new_group_upper_spec <= new_group_lower_spec:
                st.error("规格上限必须大于规格下限！")
            elif new_group_cpk_upper <= new_group_cpk_lower:
                st.error("CPK上限必须大于CPK下限！")
            else:
                new_group = {
                    'name': new_group_name,
                    'params': {
                        'center': new_group_center,
                        'upper_spec': new_group_upper_spec,
                        'lower_spec': new_group_lower_spec,
                        'cpk_upper': new_group_cpk_upper,
                        'cpk_lower': new_group_cpk_lower,
                        'mean': new_group_mean,
                        'precision': new_group_precision,
                        'count': new_group_count
                    },
                    'numbers': None,
                    'stats': None
                }
                st.session_state.condition_groups.append(new_group)
                st.session_state.current_group_index = len(st.session_state.condition_groups) - 1
                st.success(f"✓ 已添加：{new_group_name}")
                st.rerun()
        
        st.markdown('</div>', unsafe_allow_html=True)
    else:
        # 无条件组时的默认参数设置
        st.markdown('<div class="sidebar-section">', unsafe_allow_html=True)
        st.markdown('<div class="sidebar-section-title">⚙️ 核心参数</div>', unsafe_allow_html=True)
        
        center = st.number_input("目标值 (Target)", value=10.0, step=0.1, format="%.2f")
        upper_spec = st.number_input("规格上限 (USL)", value=12.0, step=0.1, format="%.2f")
        lower_spec = st.number_input("规格下限 (LSL)", value=8.0, step=0.1, format="%.2f")
        cpk_upper = st.number_input("CPK上限", value=1.67, step=0.01, format="%.2f")
        cpk_lower = st.number_input("CPK下限", value=1.33, step=0.01, format="%.2f")
        mean = st.number_input("均值 (Mean)", value=10.0, step=0.1, format="%.2f")
        st.markdown('</div>', unsafe_allow_html=True)
        
        # 其他参数
        st.markdown('<div class="sidebar-section">', unsafe_allow_html=True)
        st.markdown('<div class="sidebar-section-title">🔧 其他参数</div>', unsafe_allow_html=True)
        
        precision = st.slider("精度 (小数位)", min_value=0, max_value=6, value=2)
        count = st.slider("样本数量", min_value=1, max_value=500, value=32)
        st.markdown('</div>', unsafe_allow_html=True)
        
        # 条件组管理
        st.markdown('<div class="sidebar-section">', unsafe_allow_html=True)
        st.markdown('<div class="sidebar-section-title">🎯 条件组管理</div>', unsafe_allow_html=True)
        
        group_name = st.text_input("条件组名称", value=f"条件组 {len(st.session_state.condition_groups) + 1}")
        
        col1, col2 = st.columns(2)
        with col1:
            if st.button("➕ 添加", use_container_width=True):
                if upper_spec <= lower_spec:
                    st.error("规格上限必须大于规格下限！")
                elif cpk_upper <= cpk_lower:
                    st.error("CPK上限必须大于CPK下限！")
                else:
                    new_group = {
                        'name': group_name,
                        'params': {
                            'center': center,
                            'upper_spec': upper_spec,
                            'lower_spec': lower_spec,
                            'cpk_upper': cpk_upper,
                            'cpk_lower': cpk_lower,
                            'mean': mean,
                            'precision': precision,
                            'count': count
                        },
                        'numbers': None,
                        'stats': None
                    }
                    st.session_state.condition_groups.append(new_group)
                    st.session_state.current_group_index = len(st.session_state.condition_groups) - 1
                    st.success(f"✓ 已添加：{group_name}")
                    st.rerun()
        
        with col2:
            if st.button("🗑️ 清空", use_container_width=True):
                if st.session_state.condition_groups:
                    st.session_state.condition_groups = []
                    st.session_state.current_group_index = -1
                    st.success("✓ 已清空所有条件组")
                    st.rerun()
        st.markdown('</div>', unsafe_allow_html=True)
        

    
    # 外部条件组导入（始终显示）
    st.markdown('<div class="sidebar-section">', unsafe_allow_html=True)
    st.markdown('<div class="sidebar-section-title">📤 外部导入</div>', unsafe_allow_html=True)
    
    # 文件上传
    uploaded_file = st.file_uploader("上传条件组文件 (CSV/Excel)", type=["csv", "xlsx"])
    
    # 预检测文件是否为加密软件自动加密的文件
    def detect_encryption_software(file_obj):
        """检测文件是否被加密软件自动加密"""
        try:
            # 读取文件的前几个字节
            file_obj.seek(0)
            header = file_obj.read(8)
            file_obj.seek(0)
            
            # 检查常见的加密软件特征
            encryption_signatures = {
                b'\x50\x4b\x03\x04': '标准ZIP/Excel格式',
                b'\xd0\xcf\x11\xe0': '旧版Excel格式(.xls)',
                b'\xef\xbb\xbf': 'UTF-8 BOM (CSV)',
                b'\xff\xfe': 'UTF-16 LE (CSV)',
            }
            
            # 检查文件头
            file_signature = header[:4]
            is_standard_format = False
            detected_format = '未知格式'
            
            for sig, desc in encryption_signatures.items():
                if header.startswith(sig):
                    is_standard_format = True
                    detected_format = desc
                    break
            
            # 如果不是标准格式，可能是加密文件
            if not is_standard_format:
                # 检查是否包含加密软件的特征
                header_str = header.hex().upper()
                
                # 常见的加密软件特征（文件扩展名或标识）
                encryption_indicators = [
                    ('加密软件', ['LOCK', 'ENC', 'CRYPT', 'SECURE']),
                    ('企业加密', ['DLP', 'EDP', 'SDP', 'CDG']),
                    ('安全软件', ['SAFE', 'GUARD', 'PROTECT']),
                ]
                
                # 检查文件名是否包含加密标识
                file_name = file_obj.name.lower()
                for enc_type, indicators in encryption_indicators:
                    for indicator in indicators:
                        if indicator.lower() in file_name:
                            return {
                                'is_encrypted': True,
                                'type': enc_type,
                                'indicator': indicator,
                                'suggestion': f'检测到{enc_type}加密标识'
                            }
                
                # 检查文件大小异常（加密文件通常比原始文件大）
                file_obj.seek(0, 2)  # 移动到文件末尾
                file_size = file_obj.tell()
                file_obj.seek(0)
                
                # 如果文件大小异常小或异常大，可能是加密文件
                if file_size < 100:  # 小于100字节
                    return {
                        'is_encrypted': True,
                        'type': '可能的加密文件',
                        'indicator': '文件大小异常',
                        'suggestion': '文件大小异常，可能是加密文件或损坏文件'
                    }
            
            return {
                'is_encrypted': False,
                'format': detected_format
            }
            
        except Exception as e:
            return {
                'is_encrypted': False,
                'error': str(e)
            }
    
    if uploaded_file is not None:
        # 首先检测文件是否被加密
        encryption_check = detect_encryption_software(uploaded_file)
        
        if encryption_check.get('is_encrypted'):
            st.error(f"❌ 检测到加密文件：{encryption_check.get('suggestion', '')}")
            st.warning("⚠️ 文件可能被企业加密软件自动加密")
            st.info("""
            **解决方案：**
            
            1. **联系IT管理员**
               - 请求临时解密文件
               - 申请导入白名单权限
            
            2. **使用替代方法**
               - 将数据复制到未加密的Excel文件
               - 使用CSV格式（加密软件通常不加密CSV）
            
            3. **手动输入**
               - 在应用中手动创建条件组
               - 参考模板格式逐个添加
            
            4. **解密后上传**
               - 在安全环境中解密文件
               - 立即上传并删除临时文件
            """)
            # 仍然尝试读取，但会显示更详细的错误
        
        try:
            # 读取文件
            file_extension = uploaded_file.name.split('.')[-1].lower()
            
            if file_extension == 'csv':
                try:
                    df = pd.read_csv(uploaded_file)
                except Exception as csv_error:
                    st.error(f"❌ CSV文件读取失败：{str(csv_error)}")
                    # 继续执行，不使用return
                    continue_processing = False
                else:
                    continue_processing = True
            elif file_extension in ['xlsx', 'xls']:
                try:
                    # 尝试使用openpyxl读取
                    df = pd.read_excel(uploaded_file, engine='openpyxl')
                    continue_processing = True
                except Exception as xlsx_error:
                    # 如果失败，尝试使用xlrd读取旧格式
                    try:
                        df = pd.read_excel(uploaded_file, engine='xlrd')
                        continue_processing = True
                    except Exception as xls_error:
                        st.error(f"❌ Excel文件读取失败：{str(xlsx_error)}")
                        
                        # 提供更详细的错误诊断
                        error_msg = str(xlsx_error).lower()
                        if 'not a zip file' in error_msg:
                            st.warning("⚠️ 检测到文件格式问题")
                            
                            # 如果之前检测到加密，显示加密相关信息
                            if encryption_check.get('is_encrypted'):
                                st.error(f"🔒 确认：文件被{encryption_check.get('type', '加密软件')}加密")
                                st.info(f"加密标识：{encryption_check.get('indicator', '未知')}")
                            else:
                                st.info("""
                                **可能的原因和解决方案：**
                                
                                1. **文件被加密或受保护**
                                   - 请取消文件的密码保护
                                   - 在Excel中：文件 → 信息 → 保护工作簿 → 取消保护
                                
                                2. **文件损坏或不完整**
                                   - 尝试在Excel中打开文件，然后另存为新的.xlsx文件
                                   - 检查文件是否完整下载或复制
                                
                                3. **文件格式不正确**
                                   - 确保文件是标准的Excel格式(.xlsx)
                                   - 如果是.xls格式，请另存为.xlsx格式
                                
                                4. **临时文件或缓存文件**
                                   - 确保上传的是原始文件，不是临时文件
                                   - 重新保存文件后再上传
                                
                                5. **企业加密软件自动加密**
                                   - 文件可能被DLP/EDP等加密软件自动加密
                                   - 请联系IT管理员解密
                                   - 或尝试使用CSV格式导入
                                
                                **建议操作：**
                                - 在Excel中打开文件
                                - 点击"文件" → "另存为" → 选择"Excel工作簿(*.xlsx)"
                                - 上传新保存的文件
                                """)
                        elif 'password' in error_msg or 'encrypted' in error_msg:
                            st.warning("⚠️ 文件可能被加密或受密码保护")
                            st.info("请取消文件的密码保护后再上传")
                        else:
                            st.info("ℹ️ 可能的原因：文件损坏、格式不正确或不是标准的Excel文件")
                        continue_processing = False
            else:
                st.error("❌ 不支持的文件格式，请上传CSV或Excel文件")
                continue_processing = False
            
            # 如果读取失败，跳过后续处理
            if not continue_processing:
                # 使用pass而不是continue，因为我们不在循环中
                pass
            else:
                # 继续处理文件数据
                # 检查必要列
                required_columns = ['name', 'center', 'upper_spec', 'lower_spec', 'cpk_upper', 'cpk_lower', 'mean', 'precision', 'count']
                if all(col in df.columns for col in required_columns):
                    # 导入条件组
                    imported_count = 0
                    for _, row in df.iterrows():
                        try:
                            # 检查数据有效性
                            if float(row['upper_spec']) > float(row['lower_spec']) and float(row['cpk_upper']) > float(row['cpk_lower']):
                                new_group = {
                                    'name': str(row['name']),
                                    'params': {
                                        'center': float(row['center']),
                                        'upper_spec': float(row['upper_spec']),
                                        'lower_spec': float(row['lower_spec']),
                                        'cpk_upper': float(row['cpk_upper']),
                                        'cpk_lower': float(row['cpk_lower']),
                                        'mean': float(row['mean']),
                                        'precision': int(row['precision']),
                                        'count': int(row['count'])
                                    },
                                    'numbers': None,
                                    'stats': None
                                }
                                st.session_state.condition_groups.append(new_group)
                                imported_count += 1
                        except Exception as row_error:
                            st.warning(f"⚠️ 跳过无效行：{str(row_error)}")
                            continue
                    
                    if imported_count > 0:
                        st.session_state.current_group_index = len(st.session_state.condition_groups) - 1
                        st.success(f"✓ 成功导入 {imported_count} 个条件组")
                        st.rerun()
                    else:
                        st.error("❌ 没有有效的条件组可以导入，请检查文件数据")
                else:
                    st.error(f"❌ 文件格式不正确，缺少必要列：{[col for col in required_columns if col not in df.columns]}")
                    # 显示文件实际列名
                    st.info(f"ℹ️ 文件实际列名：{list(df.columns)}")
        except Exception as e:
            st.error(f"❌ 导入失败：{str(e)}")
            st.info("ℹ️ 请检查文件是否完整、格式是否正确")
    
    # 剪贴板粘贴功能（绕过加密文件限制）
    st.markdown("---")
    st.markdown("**📋 或者从剪贴板粘贴数据（绕过加密限制）**")
    
    clipboard_text = st.text_area(
        "粘贴Excel/CSV数据（支持从Excel直接复制粘贴）",
        height=150,
        placeholder="""从Excel复制数据，格式如下：
name\tcenter\tupper_spec\tlower_spec\tcpk_upper\tcpk_lower\tmean\tprecision\tcount
条件组1\t10.0\t12.0\t8.0\t1.67\t1.33\t10.0\t2\t32
条件组2\t20.0\t22.0\t18.0\t1.67\t1.33\t20.0\t2\t32""",
        help="从Excel复制数据后直接粘贴到这里，支持Tab分隔或逗号分隔格式"
    )
    
    if st.button("📥 从剪贴板导入", use_container_width=True, type="primary"):
        if clipboard_text.strip():
            try:
                from io import StringIO
                import re
                
                # 尝试多种方式解析数据
                df_clipboard = None
                
                # 方法1：尝试作为TSV解析（Tab分隔 - Excel默认格式）
                try:
                    df_clipboard = pd.read_csv(StringIO(clipboard_text), sep='\t')
                except:
                    pass
                
                # 方法2：尝试作为CSV解析（逗号分隔）
                if df_clipboard is None:
                    try:
                        df_clipboard = pd.read_csv(StringIO(clipboard_text), sep=',')
                    except:
                        pass
                
                # 方法3：尝试自动检测分隔符
                if df_clipboard is None:
                    try:
                        df_clipboard = pd.read_csv(StringIO(clipboard_text), sep=None, engine='python')
                    except:
                        pass
                
                # 方法4：处理空格分隔的情况（多个空格作为分隔符）
                if df_clipboard is None or len(df_clipboard.columns) == 1:
                    try:
                        # 使用正则表达式处理多个空格作为分隔符
                        df_clipboard = pd.read_csv(StringIO(clipboard_text), sep=r'\s+', engine='python')
                    except:
                        pass
                
                # 方法5：手动解析（如果上述方法都失败）
                if df_clipboard is None or len(df_clipboard.columns) < 9:
                    try:
                        lines = clipboard_text.strip().split('\n')
                        if len(lines) >= 2:
                            # 尝试用多种分隔符分割第一行（表头）
                            header = None
                            for sep in ['\t', ',', ';', '  ', ' ']:
                                if sep in lines[0]:
                                    header = [h.strip() for h in lines[0].split(sep) if h.strip()]
                                    if len(header) >= 9:
                                        break
                            
                            if header and len(header) >= 9:
                                data_rows = []
                                for line in lines[1:]:
                                    for sep in ['\t', ',', ';', '  ', ' ']:
                                        if sep in line:
                                            values = [v.strip() for v in line.split(sep) if v.strip()]
                                            if len(values) >= 9:
                                                data_rows.append(values[:9])
                                                break
                                
                                if data_rows:
                                    df_clipboard = pd.DataFrame(data_rows, columns=header[:9])
                    except Exception as manual_error:
                        st.warning(f"手动解析尝试失败：{manual_error}")
                
                if df_clipboard is not None and not df_clipboard.empty:
                    # 调试信息（开发时使用）
                    # st.write(f"检测到的列数：{len(df_clipboard.columns)}")
                    # st.write(f"列名：{list(df_clipboard.columns)}")
                    
                    # 检查必要列
                    required_columns = ['name', 'center', 'upper_spec', 'lower_spec', 'cpk_upper', 'cpk_lower', 'mean', 'precision', 'count']
                    
                    # 列名标准化（去除空格，转换为小写）
                    df_clipboard.columns = [str(col).strip().lower() for col in df_clipboard.columns]
                    
                    # 特殊处理：如果只有一个列且包含所有需要的列名（用\t分隔）
                    if len(df_clipboard.columns) == 1:
                        col_name = str(df_clipboard.columns[0])
                        if '\t' in col_name and all(req in col_name for req in required_columns):
                            # 重新解析，使用第一行作为列名
                            lines = clipboard_text.strip().split('\n')
                            if len(lines) >= 2:
                                header = [h.strip() for h in lines[0].split('\t') if h.strip()]
                                data_rows = []
                                for line in lines[1:]:
                                    values = [v.strip() for v in line.split('\t') if v.strip()]
                                    if len(values) >= len(header):
                                        data_rows.append(values[:len(header)])
                                if data_rows:
                                    df_clipboard = pd.DataFrame(data_rows, columns=header)
                                    df_clipboard.columns = [str(col).strip().lower() for col in df_clipboard.columns]
                    
                    # 再次检查是否所有必要列都存在
                    if all(col in df_clipboard.columns for col in required_columns):
                        # 导入条件组
                        imported_count = 0
                        for _, row in df_clipboard.iterrows():
                            try:
                                # 检查数据有效性
                                upper_spec = float(row['upper_spec'])
                                lower_spec = float(row['lower_spec'])
                                cpk_upper = float(row['cpk_upper'])
                                cpk_lower = float(row['cpk_lower'])
                                
                                if upper_spec > lower_spec and cpk_upper > cpk_lower:
                                    new_group = {
                                        'name': str(row['name']),
                                        'params': {
                                            'center': float(row['center']),
                                            'upper_spec': upper_spec,
                                            'lower_spec': lower_spec,
                                            'cpk_upper': cpk_upper,
                                            'cpk_lower': cpk_lower,
                                            'mean': float(row['mean']),
                                            'precision': int(row['precision']),
                                            'count': int(row['count'])
                                        },
                                        'numbers': None,
                                        'stats': None
                                    }
                                    st.session_state.condition_groups.append(new_group)
                                    imported_count += 1
                                else:
                                    st.warning(f"⚠️ 跳过无效行：{row['name']} - 规格或CPK设置不合理")
                            except Exception as row_error:
                                st.warning(f"⚠️ 跳过无效行：{str(row_error)}")
                                continue
                        
                        if imported_count > 0:
                            st.session_state.current_group_index = len(st.session_state.condition_groups) - 1
                            st.success(f"✓ 成功从剪贴板导入 {imported_count} 个条件组")
                            st.rerun()
                        else:
                            st.error("❌ 没有有效的条件组可以导入，请检查数据格式")
                    else:
                        missing_cols = [col for col in required_columns if col not in df_clipboard.columns]
                        st.error(f"❌ 数据格式不正确，缺少必要列：{missing_cols}")
                        st.info(f"ℹ️ 检测到的列名：{list(df_clipboard.columns)}")
                        st.info("""
                        **正确的数据格式：**
                        第一行必须是列名：name, center, upper_spec, lower_spec, cpk_upper, cpk_lower, mean, precision, count
                        从Excel复制时，请确保包含表头行
                        """)
                else:
                    st.error("❌ 无法解析剪贴板数据，请检查格式是否正确")
            except Exception as parse_error:
                st.error(f"❌ 解析失败：{str(parse_error)}")
                st.info("""
                **使用方法：**
                1. 在Excel中选中数据（包括表头）
                2. 按 Ctrl+C 复制
                3. 在上方文本框中按 Ctrl+V 粘贴
                4. 点击"从剪贴板导入"按钮
                """)
        else:
            st.warning("⚠️ 请先粘贴数据到文本框中")
    
    # 下载模板
    if st.button("📥 下载条件组模板"):
        # 创建模板DataFrame
        template_data = {
            'name': ['条件组1', '条件组2'],
            'center': [10.0, 20.0],
            'upper_spec': [12.0, 22.0],
            'lower_spec': [8.0, 18.0],
            'cpk_upper': [1.67, 1.67],
            'cpk_lower': [1.33, 1.33],
            'mean': [10.0, 20.0],
            'precision': [2, 2],
            'count': [32, 32]
        }
        template_df = pd.DataFrame(template_data)
        
        # 生成Excel文件
        output = io.BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            template_df.to_excel(writer, sheet_name='条件组模板', index=False)
        output.seek(0)
        
        # 提供下载
        st.download_button(
            label="下载Excel模板",
            data=output,
            file_name="条件组模板.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True
        )
    
    st.markdown('</div>', unsafe_allow_html=True)
    
    st.markdown('</div>', unsafe_allow_html=True)

# ==================== 主内容区域 ====================
if not st.session_state.condition_groups:
    # 空状态
    st.markdown("""
    <div class="empty-state">
        <div class="empty-state-icon">📋</div>
        <h3>欢迎使用 NDG</h3>
        <p>请在左侧边栏添加条件组，开始生成正态分布随机数</p>
    </div>
    """, unsafe_allow_html=True)
else:
    # 条件组选择标签
    st.markdown('<div class="control-panel">', unsafe_allow_html=True)
    st.markdown('<div class="panel-title">📑 条件组选择</div>', unsafe_allow_html=True)
    
    # 使用复选框实现多条件组选择
    col_count = min(4, len(st.session_state.condition_groups))
    cols = st.columns(col_count)
    
    # 临时存储用户的选择
    temp_selected = []
    
    for i, group in enumerate(st.session_state.condition_groups):
        col = cols[i % col_count]
        has_data = '✓' if group['numbers'] is not None else '○'
        is_selected = i in st.session_state.selected_groups
        
        with col:
            if st.checkbox(f"{has_data} {group['name']}", value=is_selected, key=f"group_{i}"):
                temp_selected.append(i)
    
    # 更新选中的条件组
    if temp_selected:
        st.session_state.selected_groups = temp_selected
        # 确保current_group_index也被更新为选中的第一个条件组
        if st.session_state.current_group_index not in temp_selected:
            st.session_state.current_group_index = temp_selected[0]
    elif st.session_state.selected_groups:
        # 如果用户取消了所有选择，保持至少一个选中
        st.error("请至少选择一个条件组")
        # 恢复之前的选择
        pass
    
    # 显示当前选中的条件组
    if st.session_state.selected_groups:
        selected_groups_info = []
        for idx in st.session_state.selected_groups:
            group = st.session_state.condition_groups[idx]
            has_data = '✓' if group['numbers'] is not None else '○'
            selected_groups_info.append(f"{group['name']} {has_data}")
        
        st.markdown(f'<div style="margin-top: 1rem; padding: 0.8rem; background: #f0f2f5; border-radius: 8px;">\n'  
                   f'<strong>已选择条件组 ({len(st.session_state.selected_groups)}):</strong>\n'  
                   f'<ul style="margin: 0.5rem 0 0 1.5rem; padding: 0;">\n'  
                   f'{"\n".join([f"<li>{info}</li>" for info in selected_groups_info])}\n'  
                   f'</ul>\n'  
                   f'<small>点击上方复选框选择/取消选择条件组</small>\n'  
                   f'</div>', unsafe_allow_html=True)
    
    # 操作按钮栏
    st.markdown('<div class="action-bar">', unsafe_allow_html=True)
    col1, col2, col3, col4 = st.columns([1, 1, 1, 1])
    
    with col1:
        if st.button("🎲 生成随机数", use_container_width=True, type="primary"):
            current_index = st.session_state.current_group_index
            group = st.session_state.condition_groups[current_index]
            params = group['params']
            
            try:
                # 验证参数类型
                params['mean'] = float(params['mean'])
                params['upper_spec'] = float(params['upper_spec'])
                params['lower_spec'] = float(params['lower_spec'])
                params['cpk_lower'] = float(params['cpk_lower'])
                params['cpk_upper'] = float(params['cpk_upper'])
                params['precision'] = int(params['precision'])
                params['count'] = int(params['count'])
                
                numbers, cp, cpk, mean_val, std_dev = generate_numbers_robust_spc(
                    params['mean'],
                    params['upper_spec'],
                    params['lower_spec'],
                    params['cpk_lower'],
                    params['cpk_upper'],
                    params['precision'],
                    params['count']
                )
                
                group['numbers'] = numbers
                group['stats'] = {
                    'cp': cp,
                    'cpk': cpk,
                    'mean': mean_val,
                    'std': std_dev
                }
                st.success(f"✓ 生成成功！CPK: {cpk:.4f}")
                st.rerun()
            except InputValidationError as e:
                st.error(f"❌ 参数错误：{str(e)}")
            except GenerationError as e:
                st.error(f"❌ 生成错误：{str(e)}")
            except StatisticsError as e:
                st.error(f"❌ 统计错误：{str(e)}")
            except ValueError as e:
                st.error(f"❌ 数值错误：{str(e)}")
            except Exception as e:
                st.error(f"❌ 未知错误：{str(e)}")
    
    with col2:
        if st.button("🔄 重新生成", use_container_width=True):
            current_index = st.session_state.current_group_index
            group = st.session_state.condition_groups[current_index]
            if group['numbers'] is not None:
                params = group['params']
                
                try:
                    # 验证参数类型
                    params['mean'] = float(params['mean'])
                    params['upper_spec'] = float(params['upper_spec'])
                    params['lower_spec'] = float(params['lower_spec'])
                    params['cpk_lower'] = float(params['cpk_lower'])
                    params['cpk_upper'] = float(params['cpk_upper'])
                    params['precision'] = int(params['precision'])
                    params['count'] = int(params['count'])
                    
                    numbers, cp, cpk, mean_val, std_dev = generate_numbers_robust_spc(
                        params['mean'],
                        params['upper_spec'],
                        params['lower_spec'],
                        params['cpk_lower'],
                        params['cpk_upper'],
                        params['precision'],
                        params['count']
                    )
                    
                    group['numbers'] = numbers
                    group['stats'] = {
                        'cp': cp,
                        'cpk': cpk,
                        'mean': mean_val,
                        'std': std_dev
                    }
                    st.success(f"✓ 重新生成成功！新CPK: {cpk:.4f}")
                    st.rerun()
                except InputValidationError as e:
                    st.error(f"❌ 参数错误：{str(e)}")
                except GenerationError as e:
                    st.error(f"❌ 生成错误：{str(e)}")
                except StatisticsError as e:
                    st.error(f"❌ 统计错误：{str(e)}")
                except ValueError as e:
                    st.error(f"❌ 数值错误：{str(e)}")
                except Exception as e:
                    st.error(f"❌ 重新生成错误：{str(e)}")
            else:
                st.warning("⚠️ 请先生成随机数")
    
    with col3:
        if st.button("🚀 批量生成全部", use_container_width=True):
            success_count = 0
            error_messages = []
            
            for i, group in enumerate(st.session_state.condition_groups):
                params = group['params']
                
                try:
                    # 验证参数类型
                    params['mean'] = float(params['mean'])
                    params['upper_spec'] = float(params['upper_spec'])
                    params['lower_spec'] = float(params['lower_spec'])
                    params['cpk_lower'] = float(params['cpk_lower'])
                    params['cpk_upper'] = float(params['cpk_upper'])
                    params['precision'] = int(params['precision'])
                    params['count'] = int(params['count'])
                    
                    numbers, cp, cpk, mean_val, std_dev = generate_numbers_robust_spc(
                        params['mean'],
                        params['upper_spec'],
                        params['lower_spec'],
                        params['cpk_lower'],
                        params['cpk_upper'],
                        params['precision'],
                        params['count']
                    )
                    
                    group['numbers'] = numbers
                    group['stats'] = {
                        'cp': cp,
                        'cpk': cpk,
                        'mean': mean_val,
                        'std': std_dev
                    }
                    success_count += 1
                except InputValidationError as e:
                    error_messages.append(f"{group['name']}: 参数错误 - {str(e)}")
                except GenerationError as e:
                    error_messages.append(f"{group['name']}: 生成错误 - {str(e)}")
                except StatisticsError as e:
                    error_messages.append(f"{group['name']}: 统计错误 - {str(e)}")
                except ValueError as e:
                    error_messages.append(f"{group['name']}: 数值错误 - {str(e)}")
                except Exception as e:
                    error_messages.append(f"{group['name']}: 未知错误 - {str(e)}")
            
            if success_count > 0:
                st.success(f"✓ 成功生成 {success_count}/{len(st.session_state.condition_groups)} 个条件组")
            if error_messages:
                for msg in error_messages[:3]:  # 只显示前3个错误
                    st.error(f"❌ {msg}")
            if success_count > 0:
                st.rerun()
    
    with col4:
        view_mode = st.radio(
            "视图模式",
            options=["单组查看", "对比视图"],
            index=0 if st.session_state.view_mode == "single" else 1,
            horizontal=True
        )
        st.session_state.view_mode = "single" if view_mode == "单组查看" else "comparison"
    
    st.markdown('</div>', unsafe_allow_html=True)
    st.markdown('</div>', unsafe_allow_html=True)
    
    # 检查是否有数据
    has_data = any(group['numbers'] is not None for group in st.session_state.condition_groups)
    
    if has_data:
        # ==================== 对比视图 ====================
        if st.session_state.view_mode == "comparison":
            st.markdown('<div class="chart-container">', unsafe_allow_html=True)
            st.markdown('<div class="chart-title">📊 多条件组对比分析</div>', unsafe_allow_html=True)
            
            # 对比表格
            comparison_data = []
            selected_groups = [st.session_state.condition_groups[idx] for idx in st.session_state.selected_groups if st.session_state.condition_groups[idx]['numbers'] is not None]
            
            for group in selected_groups:
                if group['numbers'] is not None:
                    params = group['params']
                    stats = group['stats']
                    numbers = group['numbers']
                    max_val, min_val, range_val, _ = calculate_additional_stats(numbers)
                    status_class, status_text = get_cpk_status(stats['cpk'])
                    
                    comparison_data.append({
                        '条件组': group['name'],
                        '样本数': len(numbers),
                        '均值': f"{stats['mean']:.4f}",
                        '标准差': f"{stats['std']:.4f}",
                        'CP': f"{stats['cp']:.4f}",
                        'CPK': f"{stats['cpk']:.4f}",
                        '质量等级': status_text,
                        '最大值': f"{max_val:.4f}",
                        '最小值': f"{min_val:.4f}",
                        '范围': f"{range_val:.4f}",
                        'USL': params['upper_spec'],
                        'LSL': params['lower_spec'],
                        'Target': params['center']
                    })
            
            if comparison_data:
                df_comparison = pd.DataFrame(comparison_data)
                
                # 使用HTML表格展示
                table_html = '<div class="scroll-container"><table class="comparison-table">'
                table_html += '<tr>' + ''.join([f'<th>{col}</th>' for col in df_comparison.columns]) + '</tr>'
                
                for _, row in df_comparison.iterrows():
                    table_html += '<tr>'
                    for col in df_comparison.columns:
                        value = row[col]
                        if col == '条件组':
                            table_html += f'<td class="group-name">{value}</td>'
                        elif col == '质量等级':
                            status_class = 'excellent' if value == '优秀' else ('good' if value == '良好' else 'poor')
                            table_html += f'<td><span class="quality-badge {status_class}">{value}</span></td>'
                        elif col == 'CPK':
                            cpk_val = float(value)
                            metric_class = 'metric-good' if cpk_val >= 1.67 else ('metric-warning' if cpk_val >= 1.33 else 'metric-danger')
                            table_html += f'<td class="{metric_class}">{value}</td>'
                        else:
                            table_html += f'<td>{value}</td>'
                    table_html += '</tr>'
                table_html += '</table></div>'
                
                st.markdown(table_html, unsafe_allow_html=True)
            else:
                st.info("ℹ️ 请选择至少一个已生成数据的条件组进行对比")
            
            st.markdown('</div>', unsafe_allow_html=True)
            
            if selected_groups:
                # 对比图表
                st.markdown('<div class="chart-container">', unsafe_allow_html=True)
                st.markdown('<div class="chart-title">📈 分布曲线对比</div>', unsafe_allow_html=True)
                
                # 图表显示模式选择
                chart_mode = st.radio(
                    "图表显示模式",
                    options=["合并显示（同一坐标轴）", "单独显示（各自坐标轴）"],
                    index=0,
                    horizontal=True
                )
                
                if chart_mode == "合并显示（同一坐标轴）":
                    # 所有组在同一个坐标轴显示
                    fig, ax = plt.subplots(figsize=(14, 6))
                    colors = ['#667eea', '#f093fb', '#4facfe', '#43e97b', '#fa709a', '#fee140']
                    
                    for i, group in enumerate(selected_groups):
                        if group['numbers'] is not None:
                            params = group['params']
                            stats = group['stats']
                            numbers = group['numbers']
                            color = colors[i % len(colors)]
                            
                            # 绘制直方图（透明度较低）
                            ax.hist(numbers, bins=15, density=True, alpha=0.3, color=color, 
                                   edgecolor=color, linewidth=1.5, label=f"{group['name']} (直方图)")
                            
                            # 绘制正态分布曲线
                            x = np.linspace(params['lower_spec'] - 0.5, params['upper_spec'] + 0.5, 100)
                            y = ((1 / (stats['std'] * np.sqrt(2 * np.pi))) * 
                                 np.exp(-0.5 * ((x - stats['mean']) / stats['std']) ** 2))
                            ax.plot(x, y, color=color, linewidth=2.5, 
                                   label=f"{group['name']} (CPK={stats['cpk']:.2f})")
                    
                    # 绘制规格线（使用第一个选中组的规格线）
                    first_group = next((g for g in selected_groups if g['numbers'] is not None), None)
                    if first_group:
                        params = first_group['params']
                        ax.axvline(x=params['lower_spec'], color='red', linestyle='--', linewidth=2, label='LSL')
                        ax.axvline(x=params['upper_spec'], color='red', linestyle='--', linewidth=2, label='USL')
                        ax.axvline(x=params['center'], color='green', linestyle='-.', linewidth=2, label='Target')
                    
                    ax.set_xlabel('数值', fontsize=12, fontweight='bold')
                    ax.set_ylabel('概率密度', fontsize=12, fontweight='bold')
                    ax.set_title('多条件组分布曲线对比', fontsize=14, fontweight='bold')
                    ax.legend(loc='upper right', fontsize=9, frameon=True, ncol=2)
                    ax.grid(True, alpha=0.3, linestyle='--')
                    
                    fig.tight_layout()
                    st.pyplot(fig)
                else:
                    # 每组单独显示，拥有各自的坐标轴
                    groups_with_data = [group for group in selected_groups if group['numbers'] is not None]
                    
                    if groups_with_data:
                        # 根据组数计算布局
                        num_groups = len(groups_with_data)
                        if num_groups <= 2:
                            rows, cols = 1, num_groups
                        elif num_groups <= 4:
                            rows, cols = 2, 2
                        else:
                            rows, cols = (num_groups + 2) // 3, 3
                        
                        fig, axes = plt.subplots(rows, cols, figsize=(4 * cols, 4 * rows))
                        axes = np.array(axes).reshape(-1)  # 确保是一维数组
                        colors = ['#667eea', '#f093fb', '#4facfe', '#43e97b', '#fa709a', '#fee140']
                        
                        for i, (ax, group) in enumerate(zip(axes, groups_with_data)):
                            params = group['params']
                            stats = group['stats']
                            numbers = group['numbers']
                            color = colors[i % len(colors)]
                            
                            # 绘制直方图
                            ax.hist(numbers, bins=10, density=False, alpha=0.7, color=color, 
                                   edgecolor='white', linewidth=1, label='直方图')
                            
                            # 绘制正态分布曲线
                            x = np.linspace(params['lower_spec'] - 0.5, params['upper_spec'] + 0.5, 100)
                            bin_width = (np.max(numbers) - np.min(numbers)) / 10 if np.max(numbers) != np.min(numbers) else 1
                            y = ((1 / (stats['std'] * np.sqrt(2 * np.pi))) * 
                                 np.exp(-0.5 * ((x - stats['mean']) / stats['std']) ** 2)) * len(numbers) * bin_width
                            ax.plot(x, y, 'r-', linewidth=2, label='正态曲线')
                            
                            # 绘制参考线
                            ax.axvline(x=params['lower_spec'], color='red', linestyle='--', linewidth=1.5, label='LSL')
                            ax.axvline(x=params['upper_spec'], color='red', linestyle='--', linewidth=1.5, label='USL')
                            if params['center'] != params['mean']:
                                ax.axvline(x=params['center'], color='green', linestyle='--', linewidth=1.5, label='Target')
                            ax.axvline(x=stats['mean'], color='blue', linestyle='-', linewidth=1.5, label='均值')
                            
                            # 设置图表属性
                            ax.set_xlabel('数值', fontsize=10)
                            ax.set_ylabel('频数', fontsize=10)
                            ax.set_title(f'{group["name"]}\nCPK={stats["cpk"]:.2f}', fontsize=11, fontweight='bold')
                            ax.tick_params(axis='both', labelsize=8)
                            
                            # 设置x轴范围
                            x_min = params['lower_spec'] - 0.5
                            x_max = params['upper_spec'] + 0.5
                            ax.set_xlim(x_min, x_max)
                            
                            # 添加网格
                            ax.grid(True, alpha=0.3, linestyle='--')
                            
                            # 优化图例
                            ax.legend(loc='upper right', fontsize=7, frameon=True, framealpha=0.9)
                        
                        # 隐藏多余的子图
                        for j in range(len(groups_with_data), len(axes)):
                            axes[j].set_visible(False)
                        
                        fig.tight_layout()
                        st.pyplot(fig)
            
            st.markdown('</div>', unsafe_allow_html=True)
            
            # 数据导出
            if comparison_data:
                st.markdown('<div class="data-section">', unsafe_allow_html=True)
                st.markdown('<div class="panel-title">💾 数据导出</div>', unsafe_allow_html=True)
                
                col1, col2 = st.columns([1, 3])
                with col1:
                    if st.button("📥 导出选中数据到Excel", use_container_width=True, type="primary"):
                        try:
                            # 辅助函数：生成分布图表并转换为图片
                            def generate_distribution_chart(group_name, numbers, params, stats):
                                # 创建图形
                                fig, (ax1, ax2) = plt.subplots(2, 1, figsize=(12, 10))
                                fig.suptitle(f'{group_name} - 数据分布分析', fontsize=16, fontweight='bold')
                                
                                # 直方图
                                ax1.hist(numbers, bins=20, alpha=0.7, color='#667eea', edgecolor='black')
                                ax1.axvline(params['center'], color='red', linestyle='--', linewidth=2, label=f'目标值: {params["center"]}')
                                ax1.axvline(params['upper_spec'], color='green', linestyle='--', linewidth=2, label=f'规格上限: {params["upper_spec"]}')
                                ax1.axvline(params['lower_spec'], color='green', linestyle='--', linewidth=2, label=f'规格下限: {params["lower_spec"]}')
                                ax1.axvline(stats['mean'], color='blue', linestyle='-', linewidth=2, label=f'实际均值: {stats["mean"]:.2f}')
                                ax1.set_xlabel('数值')
                                ax1.set_ylabel('频数')
                                ax1.set_title('数据分布直方图')
                                ax1.legend()
                                ax1.grid(True, alpha=0.3)
                                
                                # 正态分布曲线
                                x_min = min(numbers) - 0.5
                                x_max = max(numbers) + 0.5
                                x_values = np.linspace(x_min, x_max, 100)
                                pdf = (1 / (stats['std'] * np.sqrt(2 * np.pi))) * np.exp(-0.5 * ((x_values - stats['mean']) / stats['std']) ** 2)
                                
                                ax2.plot(x_values, pdf, 'r-', linewidth=2, label='正态分布曲线')
                                ax2.axvline(params['center'], color='red', linestyle='--', linewidth=2, label=f'目标值: {params["center"]}')
                                ax2.axvline(params['upper_spec'], color='green', linestyle='--', linewidth=2, label=f'规格上限: {params["upper_spec"]}')
                                ax2.axvline(params['lower_spec'], color='green', linestyle='--', linewidth=2, label=f'规格下限: {params["lower_spec"]}')
                                ax2.axvline(stats['mean'], color='blue', linestyle='-', linewidth=2, label=f'实际均值: {stats["mean"]:.2f}')
                                ax2.set_xlabel('数值')
                                ax2.set_ylabel('概率密度')
                                ax2.set_title('正态分布曲线图')
                                ax2.legend()
                                ax2.grid(True, alpha=0.3)
                                
                                # 调整布局
                                plt.tight_layout(rect=[0, 0, 1, 0.95])
                                
                                # 转换为BytesIO对象
                                buf = BytesIO()
                                plt.savefig(buf, format='png', dpi=150)
                                buf.seek(0)
                                plt.close()
                                
                                return buf
                            
                            output = io.BytesIO()
                            wb = Workbook()
                            
                            # 定义样式
                            title_font = Font(name='微软雅黑', size=18, bold=True, color='1F4E78')
                            subtitle_font = Font(name='微软雅黑', size=14, bold=True, color='1F4E78')
                            header_font = Font(name='微软雅黑', size=11, bold=True, color='FFFFFF')
                            normal_font = Font(name='微软雅黑', size=10)
                            bold_font = Font(name='微软雅黑', size=10, bold=True)
                            
                            header_fill = PatternFill(start_color='1F4E78', end_color='1F4E78', fill_type='solid')
                            light_fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
                            alternate_fill = PatternFill(start_color='F0F4F8', end_color='F0F4F8', fill_type='solid')
                            green_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
                            red_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
                            yellow_fill = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
                            
                            thin_border = Border(
                                left=Side(style='thin', color='B4B4B4'),
                                right=Side(style='thin', color='B4B4B4'),
                                top=Side(style='thin', color='B4B4B4'),
                                bottom=Side(style='thin', color='B4B4B4')
                            )
                            
                            thick_border = Border(
                                left=Side(style='thick', color='1F4E78'),
                                right=Side(style='thick', color='1F4E78'),
                                top=Side(style='thick', color='1F4E78'),
                                bottom=Side(style='thick', color='1F4E78')
                            )
                            
                            center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
                            left_align = Alignment(horizontal='left', vertical='center', wrap_text=True)
                            right_align = Alignment(horizontal='right', vertical='center', wrap_text=True)
                            
                            # ========== 1. 所有条件组的统计结果对比汇总表 ==========
                            ws_summary = wb.active
                            ws_summary.title = '统计结果对比汇总'
                            
                            # 汇总数据
                            summary_df = pd.DataFrame(comparison_data)
                            
                            # 主标题
                            ws_summary.merge_cells('A1:M1')
                            ws_summary['A1'] = '统计结果对比汇总表'
                            ws_summary['A1'].font = title_font
                            ws_summary['A1'].alignment = center_align
                            ws_summary.row_dimensions[1].height = 45
                            
                            # 表头
                            headers = ['条件组', '样本数', '均值', '标准差', 'CP', 'CPK', '质量等级', '最大值', '最小值', '范围', 'USL', 'LSL', 'Target']
                            for col, header in enumerate(headers, 1):
                                cell = ws_summary.cell(3, col, header)
                                cell.font = header_font
                                cell.fill = header_fill
                                cell.alignment = center_align
                                cell.border = thin_border
                            ws_summary.row_dimensions[3].height = 30
                            
                            # 数据
                            for row_idx, row in enumerate(summary_df.itertuples(), 4):
                                row_data = list(row)[1:]  # 跳过索引
                                
                                for col_idx, value in enumerate(row_data, 1):
                                    cell = ws_summary.cell(row_idx, col_idx, value)
                                    cell.font = normal_font
                                    cell.border = thin_border
                                    
                                    # 根据列类型设置对齐方式
                                    if col_idx == 1:  # 条件组名称
                                        cell.alignment = left_align
                                    else:  # 数值类型
                                        cell.alignment = right_align
                                    
                                    # 质量等级着色
                                    if col_idx == 7:
                                        status = value
                                        if status == '优秀':
                                            cell.fill = green_fill
                                            cell.font = Font(name='微软雅黑', size=10, bold=True, color='006100')
                                            cell.alignment = center_align
                                        elif status == '良好':
                                            cell.fill = yellow_fill
                                            cell.font = Font(name='微软雅黑', size=10, bold=True, color='856404')
                                            cell.alignment = center_align
                                        else:
                                            cell.fill = red_fill
                                            cell.font = Font(name='微软雅黑', size=10, bold=True, color='9C0006')
                                            cell.alignment = center_align
                                    else:
                                        # 交替行颜色
                                        if row_idx % 2 == 0:
                                            cell.fill = light_fill
                                        else:
                                            cell.fill = alternate_fill
                            
                            # 设置列宽
                            ws_summary.column_dimensions['A'].width = 18  # 条件组名称
                            ws_summary.column_dimensions['B'].width = 10  # 样本数
                            ws_summary.column_dimensions['C'].width = 12  # 均值
                            ws_summary.column_dimensions['D'].width = 12  # 标准差
                            ws_summary.column_dimensions['E'].width = 10  # CP
                            ws_summary.column_dimensions['F'].width = 10  # CPK
                            ws_summary.column_dimensions['G'].width = 12  # 质量等级
                            ws_summary.column_dimensions['H'].width = 12  # 最大值
                            ws_summary.column_dimensions['I'].width = 12  # 最小值
                            ws_summary.column_dimensions['J'].width = 12  # 范围
                            ws_summary.column_dimensions['K'].width = 10  # USL
                            ws_summary.column_dimensions['L'].width = 10  # LSL
                            ws_summary.column_dimensions['M'].width = 10  # Target
                            
                            # 冻结表头
                            ws_summary.freeze_panes = 'A4'
                            
                            # ========== 2. 所有条件组的随机数汇总表 ==========
                            ws_combined = wb.create_sheet('随机数汇总表')
                            
                            # 主标题
                            ws_combined.merge_cells('A1:Z1')
                            ws_combined['A1'] = '随机数汇总表'
                            ws_combined['A1'].font = title_font
                            ws_combined['A1'].alignment = center_align
                            ws_combined.row_dimensions[1].height = 45
                            
                            # 添加副标题
                            ws_combined.merge_cells('A2:Z2')
                            ws_combined['A2'] = f'参数对比分析 - 生成时间: {datetime.now().strftime("%Y-%m-%d %H:%M:%S")}'
                            ws_combined['A2'].font = subtitle_font
                            ws_combined['A2'].alignment = center_align
                            ws_combined.row_dimensions[2].height = 30
                            
                            # 参数类型列表
                            params_types = ['目标值(T)', '规格上限(USL)', '规格下限(LSL)', 'CPK下限', 'CPK上限', '均值(Mean)', '样本数(N)']
                            
                            # 写入左侧参数类型列
                            for row_idx, param_type in enumerate(params_types, 4):
                                cell = ws_combined.cell(row_idx, 1, param_type)
                                cell.font = header_font
                                cell.fill = header_fill
                                cell.alignment = center_align
                                cell.border = thin_border
                            ws_combined.row_dimensions[3].height = 20  # 空行
                            
                            # 为每个条件组创建数据列
                            for col_idx, group in enumerate(selected_groups, 2):
                                params = group['params']
                                numbers = group['numbers']
                                
                                # 条件组名称
                                cell = ws_combined.cell(3, col_idx, group['name'])
                                cell.font = subtitle_font
                                cell.alignment = center_align
                                cell.border = thin_border
                                cell.fill = light_fill
                                ws_combined.row_dimensions[3].height = 30
                                
                                # 写入参数值
                                params_values = [params['center'], params['upper_spec'], params['lower_spec'], 
                                               params['cpk_lower'], params['cpk_upper'], params['mean'], params['count']]
                                
                                for row_idx, param_value in enumerate(params_values, 4):
                                    cell = ws_combined.cell(row_idx, col_idx, param_value)
                                    cell.font = normal_font
                                    cell.border = thin_border
                                    cell.alignment = right_align
                                    
                                    # 交替行颜色
                                    if (row_idx - 3) % 2 == 0:
                                        cell.fill = alternate_fill
                                    else:
                                        cell.fill = light_fill
                                
                                # 写入随机数数据
                                if numbers is not None:
                                    # 随机数表头
                                    random_start_row = len(params_types) + 4
                                    cell = ws_combined.cell(random_start_row, col_idx, '随机数')
                                    cell.font = header_font
                                    cell.fill = header_fill
                                    cell.alignment = center_align
                                    cell.border = thin_border
                                    ws_combined.row_dimensions[random_start_row].height = 25
                                    
                                    # 写入随机数
                                    for row_idx, num in enumerate(numbers, random_start_row + 1):
                                        cell = ws_combined.cell(row_idx, col_idx, round(num, params['precision']))
                                        cell.font = normal_font
                                        cell.border = thin_border
                                        cell.alignment = right_align
                                        
                                        # 交替行颜色
                                        if (row_idx - random_start_row) % 2 == 0:
                                            cell.fill = light_fill
                                        else:
                                            cell.fill = alternate_fill
                            
                            # 设置列宽
                            ws_combined.column_dimensions['A'].width = 18  # 参数类型列
                            
                            # 设置每个条件组的列宽
                            for col in range(2, len(selected_groups) + 2):
                                col_letter = get_column_letter(col)
                                ws_combined.column_dimensions[col_letter].width = 16
                            
                            # 冻结窗格
                            ws_combined.freeze_panes = 'B4'
                            
                            # ========== 3. 每个条件组的随机数和分布曲线图 ==========
                            for group in selected_groups:
                                if group['numbers'] is not None:
                                    params = group['params']
                                    stats = group['stats']
                                    numbers = group['numbers']
                                    
                                    # 创建条件组工作表
                                    sheet_name = f'{group["name"]}_数据'
                                    if len(sheet_name) > 31:
                                        sheet_name = sheet_name[:31]
                                    
                                    # 原始数据
                                    ws_data = wb.create_sheet(sheet_name)
                                    
                                    # 主标题
                                    ws_data.merge_cells('A1:G1')
                                    ws_data['A1'] = f'{group["name"]} - 随机数数据和分布分析'
                                    ws_data['A1'].font = title_font
                                    ws_data['A1'].alignment = center_align
                                    ws_data.row_dimensions[1].height = 45
                                    
                                    # 添加副标题
                                    ws_data.merge_cells('A2:G2')
                                    ws_data['A2'] = f'生成时间: {datetime.now().strftime("%Y-%m-%d %H:%M:%S")}'
                                    ws_data['A2'].font = subtitle_font
                                    ws_data['A2'].alignment = center_align
                                    ws_data.row_dimensions[2].height = 30
                                    
                                    # 添加参数摘要表
                                    ws_data.merge_cells('A4:B4')
                                    ws_data['A4'] = '参数设置'
                                    ws_data['A4'].font = subtitle_font
                                    ws_data['A4'].alignment = center_align
                                    ws_data['A4'].fill = light_fill
                                    ws_data['A4'].border = thin_border
                                    
                                    # 参数列表
                                    param_list = [
                                        ('目标值(T)', params['center']),
                                        ('规格上限(USL)', params['upper_spec']),
                                        ('规格下限(LSL)', params['lower_spec']),
                                        ('CPK下限', params['cpk_lower']),
                                        ('CPK上限', params['cpk_upper']),
                                        ('均值(Mean)', params['mean']),
                                        ('样本数(N)', params['count']),
                                        ('小数精度', params['precision'])
                                    ]
                                    
                                    # 写入参数
                                    for row_idx, (param_name, param_value) in enumerate(param_list, 5):
                                        # 参数名称
                                        cell = ws_data.cell(row_idx, 1, param_name)
                                        cell.font = header_font
                                        cell.fill = header_fill
                                        cell.alignment = center_align
                                        cell.border = thin_border
                                        
                                        # 参数值
                                        cell = ws_data.cell(row_idx, 2, param_value)
                                        cell.font = normal_font
                                        cell.alignment = right_align
                                        cell.border = thin_border
                                        
                                        # 交替行颜色
                                        if (row_idx - 4) % 2 == 0:
                                            cell.fill = alternate_fill
                                        else:
                                            cell.fill = light_fill
                                    
                                    # 添加统计结果摘要
                                    ws_data.merge_cells('C4:G4')
                                    ws_data['C4'] = '统计结果'
                                    ws_data['C4'].font = subtitle_font
                                    ws_data['C4'].alignment = center_align
                                    ws_data['C4'].fill = light_fill
                                    ws_data['C4'].border = thin_border
                                    
                                    # 计算额外统计信息
                                    max_val, min_val, range_val, _ = calculate_additional_stats(numbers)
                                    
                                    # 统计结果列表
                                    stats_list = [
                                        ('均值', stats['mean']),
                                        ('标准差', stats['std']),
                                        ('CP', stats['cp']),
                                        ('CPK', stats['cpk']),
                                        ('最大值', max_val),
                                        ('最小值', min_val),
                                        ('范围', range_val)
                                    ]
                                    
                                    # 写入统计结果
                                    for row_idx, (stat_name, stat_value) in enumerate(stats_list, 5):
                                        cell = ws_data.cell(row_idx, 3, stat_name)
                                        cell.font = header_font
                                        cell.fill = header_fill
                                        cell.alignment = center_align
                                        cell.border = thin_border
                                        
                                        cell = ws_data.cell(row_idx, 4, round(stat_value, 4))
                                        cell.font = normal_font
                                        cell.alignment = right_align
                                        cell.border = thin_border
                                        
                                        # 交替行颜色
                                        if (row_idx - 4) % 2 == 0:
                                            cell.fill = alternate_fill
                                        else:
                                            cell.fill = light_fill
                                    
                                    # 生成分布图表并插入
                                    chart_buf = generate_distribution_chart(group['name'], numbers, params, stats)
                                    img = Image(chart_buf)
                                    
                                    # 调整图片大小，使其更适合工作表
                                    img.width = 700
                                    img.height = 450
                                    
                                    # 插入图片到更合理的位置
                                    ws_data.add_image(img, 'A14')
                                    
                                    # 设置行高以容纳图片
                                    ws_data.row_dimensions[14].height = 280
                                    
                                    # 添加随机数数据
                                    random_data_start = 25
                                    ws_data.merge_cells(f'A{random_data_start}:G{random_data_start}')
                                    ws_data[f'A{random_data_start}'] = '随机数数据'
                                    ws_data[f'A{random_data_start}'].font = subtitle_font
                                    ws_data[f'A{random_data_start}'].alignment = center_align
                                    ws_data[f'A{random_data_start}'].fill = light_fill
                                    ws_data[f'A{random_data_start}'].border = thin_border
                                    
                                    # 表头
                                    headers = ['序号', '随机数']
                                    for col, header in enumerate(headers, 1):
                                        cell = ws_data.cell(random_data_start + 1, col, header)
                                        cell.font = header_font
                                        cell.fill = header_fill
                                        cell.alignment = center_align
                                        cell.border = thin_border
                                    
                                    # 数据
                                    for idx, num in enumerate(numbers, 1):
                                        row_data = [idx, round(num, params['precision'])]
                                        
                                        for col_idx, value in enumerate(row_data, 1):
                                            cell = ws_data.cell(idx + random_data_start + 1, col_idx, value)
                                            cell.font = normal_font
                                            cell.border = thin_border
                                            cell.alignment = center_align
                                            
                                            # 交替行颜色
                                            if (idx + random_data_start + 1) % 2 == 0:
                                                cell.fill = light_fill
                                            else:
                                                cell.fill = alternate_fill
                                    
                                    # 设置列宽
                                    ws_data.column_dimensions['A'].width = 12  # 序号/参数名称
                                    ws_data.column_dimensions['B'].width = 18  # 随机数/参数值
                                    ws_data.column_dimensions['C'].width = 12  # 统计名称
                                    ws_data.column_dimensions['D'].width = 15  # 统计值
                                    ws_data.column_dimensions['E'].width = 5   # 间距
                                    ws_data.column_dimensions['F'].width = 5   # 间距
                                    ws_data.column_dimensions['G'].width = 5   # 间距
                                    
                                    # 取消冻结窗格
                                    ws_data.freeze_panes = None
                            
                            # 保存工作簿
                            wb.save(output)
                            output.seek(0)
                            
                            # 使用st.download_button替代自定义HTML链接
                            st.download_button(
                                label="点击下载 Excel 文件",
                                data=output,
                                file_name=f"NDG_导出报告_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
                                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                use_container_width=True
                            )
                            st.success("✓ Excel文件已生成：包含统计结果对比汇总、所有条件组随机数汇总表和每个条件组的随机数及分布曲线图")
                            
                        except Exception as e:
                            st.error(f"导出失败：{str(e)}")
                            st.exception(e)
                        finally:
                            st.markdown('</div>', unsafe_allow_html=True)
        
        # ==================== 单组查看视图 ====================
        else:
            current_index = st.session_state.current_group_index
            group = st.session_state.condition_groups[current_index]
            
            if group['numbers'] is not None:
                params = group['params']
                stats = group['stats']
                numbers = group['numbers']
                max_val, min_val, range_val, variance = calculate_additional_stats(numbers)
                normality_result = perform_normality_tests(numbers)
                status_class, status_text = get_cpk_status(stats['cpk'])
                cpk_color = get_cpk_color(stats['cpk'])
                
                # 统计指标面板 - 横向紧凑布局
                st.markdown('<div class="control-panel">', unsafe_allow_html=True)
                st.markdown(f'<div class="panel-title">📊 {group["name"]} - 统计指标</div>', unsafe_allow_html=True)
                
                # 质量等级徽章
                st.markdown(f'<div style="margin-bottom: 1rem;"><span class="quality-badge {status_class}">质量等级：{status_text}</span></div>', unsafe_allow_html=True)
                
                # 横向统计指标布局 - 使用单个HTML字符串确保flex布局生效
                stats_row_html = '<div style="display: flex; flex-wrap: wrap; gap: 0.5rem; align-items: center;">' 
                
                # 基础统计指标
                stats_display = [
                    ("样本数量", f"{len(numbers)}", ""),
                    ("均值", f"{stats['mean']:.4f}", ""),
                    ("标准差", f"{stats['std']:.4f}", ""),
                    ("CP", f"{stats['cp']:.4f}", ""),
                    ("CPK", f"{stats['cpk']:.4f}", cpk_color),
                    ("最大值", f"{max_val:.4f}", ""),
                    ("最小值", f"{min_val:.4f}", ""),
                    ("范围", f"{range_val:.4f}", ""),
                ]
                
                # 正态性检验指标
                normality_color = "#27ae60" if normality_result['is_normal'] else "#e74c3c"
                normality_text = "✓" if normality_result['is_normal'] else "✗"
                
                stats_display.extend([
                    ("正态性", f"{normality_text} {'符合' if normality_result['is_normal'] else '不符合'}", normality_color),
                    ("Shapiro p值", f"{normality_result['shapiro_p']:.4f}", ""),
                    ("偏度", f"{normality_result['skewness']:.4f}", ""),
                    ("峰度", f"{normality_result['kurtosis']:.4f}", ""),
                ])
                
                for label, value, color in stats_display:
                    value_style = f"color: {color}; font-weight: 600;" if color else "font-weight: 600;"
                    stats_row_html += f'<div style="display: flex; flex-direction: column; align-items: center; padding: 0.4rem 0.8rem; background: #f8f9fa; border-radius: 6px; min-width: 70px; font-size: 0.9rem;"><div style="font-size: 0.7rem; color: #666; text-align: center;">{label}</div><div style="{value_style}">{value}</div></div>' 
                
                stats_row_html += '</div>'
                st.markdown(stats_row_html, unsafe_allow_html=True)
                
                st.markdown('</div>', unsafe_allow_html=True)
                
                # 分布曲线
                st.markdown('<div class="chart-container">', unsafe_allow_html=True)
                st.markdown('<div class="chart-title">分布曲线</div>', unsafe_allow_html=True)
                
                fig, ax = plt.subplots(figsize=(12, 6))
                ax.set_facecolor('#fafafa')
                
                # 绘制直方图
                n, bins, patches = ax.hist(numbers, bins=15, density=False, 
                                       alpha=0.7, color='#667eea', 
                                       edgecolor='white', linewidth=1, label='直方图')
                
                # 正态分布曲线
                x = np.linspace(params['lower_spec'] - 0.5, params['upper_spec'] + 0.5, 200)
                bin_width = bins[1] - bins[0] if len(bins) > 1 else (max_val - min_val) / 15
                y = ((1 / (stats['std'] * np.sqrt(2 * np.pi))) * 
                     np.exp(-0.5 * ((x - stats['mean']) / stats['std']) ** 2)) * len(numbers) * bin_width
                
                ax.plot(x, y, color='#e74c3c', linewidth=3, label='正态分布曲线')
                
                # 参考线
                ax.axvline(x=params['lower_spec'], color='#e74c3c', linestyle='--', linewidth=2, label=f'LSL ({params["lower_spec"]})')
                ax.axvline(x=params['upper_spec'], color='#e74c3c', linestyle='--', linewidth=2, label=f'USL ({params["upper_spec"]})')
                ax.axvline(x=params['center'], color='#27ae60', linestyle='-.', linewidth=2, label=f'Target ({params["center"]})')
                ax.axvline(x=stats['mean'], color='#3498db', linestyle='-', linewidth=2, label=f'均值 ({stats["mean"]:.4f})')
                
                # 填充规格限区域
                ax.axvspan(params['lower_spec'] - 1, params['lower_spec'], alpha=0.1, color='red', label='不合格区')
                ax.axvspan(params['upper_spec'], params['upper_spec'] + 1, alpha=0.1, color='red')
                
                ax.set_xlabel('数值', fontsize=12, fontweight='bold')
                ax.set_ylabel('频数', fontsize=12, fontweight='bold')
                ax.set_title(f'{group["name"]} - 分布曲线 (CPK={stats["cpk"]:.4f})', fontsize=14, fontweight='bold')
                ax.legend(loc='upper right', fontsize=9, frameon=True, ncol=2)
                ax.grid(True, alpha=0.3, linestyle='--')
                
                fig.tight_layout()
                st.pyplot(fig)
                st.markdown('</div>', unsafe_allow_html=True)
                
                # 随机数数据表
                st.markdown('<div class="data-section">', unsafe_allow_html=True)
                st.markdown('<div class="panel-title">📝 生成的随机数数据</div>', unsafe_allow_html=True)
                
                numbers_df = pd.DataFrame({
                    '样本ID': [f"S{j+1}" for j in range(len(numbers))],
                    '数值': [round(num, params['precision']) for num in numbers]
                })
                
                col1, col2 = st.columns([3, 1])
                with col1:
                    st.dataframe(numbers_df, use_container_width=True, height=400)
                
                with col2:
                    st.markdown("**数据导出**")
                    
                    # 导出当前组Excel - 质量经理专业报告格式
                    if st.button("📥 导出Excel", use_container_width=True):
                        output = io.BytesIO()
                        
                        # 使用openpyxl引擎以支持样式和图表
                        from openpyxl import Workbook
                        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
                        from openpyxl.chart import LineChart, BarChart, Reference
                        from openpyxl.chart.label import DataLabelList
                        from openpyxl.utils.dataframe import dataframe_to_rows
                        from openpyxl.drawing.image import Image as XLImage
                        
                        wb = Workbook()
                        
                        # 定义样式
                        title_font = Font(name='微软雅黑', size=18, bold=True, color='1F4E78')
                        subtitle_font = Font(name='微软雅黑', size=14, bold=True, color='1F4E78')
                        header_font = Font(name='微软雅黑', size=11, bold=True, color='FFFFFF')
                        normal_font = Font(name='微软雅黑', size=10)
                        bold_font = Font(name='微软雅黑', size=10, bold=True)
                        small_font = Font(name='微软雅黑', size=9)
                        
                        header_fill = PatternFill(start_color='1F4E78', end_color='1F4E78', fill_type='solid')
                        subheader_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
                        light_fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
                        green_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
                        red_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
                        yellow_fill = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
                        orange_fill = PatternFill(start_color='F4B084', end_color='F4B084', fill_type='solid')
                        
                        thin_border = Border(
                            left=Side(style='thin', color='B4B4B4'),
                            right=Side(style='thin', color='B4B4B4'),
                            top=Side(style='thin', color='B4B4B4'),
                            bottom=Side(style='thin', color='B4B4B4')
                        )
                        
                        center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
                        left_align = Alignment(horizontal='left', vertical='center', wrap_text=True)
                        right_align = Alignment(horizontal='right', vertical='center')
                        
                        # 获取CPK等级和颜色
                        def get_cpk_level_info(cpk_value):
                            if cpk_value >= 2.0:
                                return 'A+级（卓越）', '卓越', '过程能力充分，无需改进', green_fill, 6
                            elif cpk_value >= 1.67:
                                return 'A级（优秀）', '优秀', '过程能力良好，维持现状', green_fill, 5
                            elif cpk_value >= 1.33:
                                return 'B级（合格）', '合格', '过程能力尚可，建议监控', yellow_fill, 4
                            elif cpk_value >= 1.0:
                                return 'C级（警告）', '警告', '过程能力不足，需要改进', orange_fill, 3
                            else:
                                return 'D级（不合格）', '不合格', '过程能力严重不足，立即整改', red_fill, 2
                        
                        cpk_level, quality_status, suggestion, cpk_fill, sigma_level = get_cpk_level_info(stats['cpk'])
                        
                        # 计算额外指标
                        # 不合格率估算（基于正态分布）
                        from scipy import stats as scipy_stats
                        z_upper = (params['upper_spec'] - stats['mean']) / stats['std'] if stats['std'] > 0 else 0
                        z_lower = (params['lower_spec'] - stats['mean']) / stats['std'] if stats['std'] > 0 else 0
                        defect_rate_upper = (1 - scipy_stats.norm.cdf(z_upper)) * 100 if stats['std'] > 0 else 0
                        defect_rate_lower = scipy_stats.norm.cdf(z_lower) * 100 if stats['std'] > 0 else 0
                        total_defect_rate = defect_rate_upper + defect_rate_lower
                        
                        # PPK计算（使用总体标准差）
                        pp = (params['upper_spec'] - params['lower_spec']) / (6 * np.std(numbers, ddof=0)) if np.std(numbers, ddof=0) > 0 else 0
                        ppk_upper = (params['upper_spec'] - stats['mean']) / (3 * np.std(numbers, ddof=0)) if np.std(numbers, ddof=0) > 0 else 0
                        ppk_lower = (stats['mean'] - params['lower_spec']) / (3 * np.std(numbers, ddof=0)) if np.std(numbers, ddof=0) > 0 else 0
                        ppk = min(ppk_upper, ppk_lower)
                        
                        # 合格率
                        within_spec = sum(1 for n in numbers if params['lower_spec'] <= n <= params['upper_spec'])
                        yield_rate = (within_spec / len(numbers)) * 100 if len(numbers) > 0 else 0
                        
                        # ========== 1. 执行摘要（质量经理视角） ==========
                        ws_cover = wb.active
                        ws_cover.title = '执行摘要'
                        
                        # 主标题
                        ws_cover.merge_cells('A1:F1')
                        ws_cover['A1'] = '过程能力分析报告'
                        ws_cover['A1'].font = title_font
                        ws_cover['A1'].alignment = center_align
                        ws_cover.row_dimensions[1].height = 45
                        
                        # 副标题
                        ws_cover.merge_cells('A2:F2')
                        ws_cover['A2'] = f'条件组：{group["name"]} | 报告编号：NDG-{datetime.now().strftime("%Y%m%d%H%M%S")}'
                        ws_cover['A2'].font = subtitle_font
                        ws_cover['A2'].alignment = center_align
                        ws_cover.row_dimensions[2].height = 30
                        
                        # 关键指标区域
                        ws_cover['A4'] = '关键质量指标（KPI）'
                        ws_cover['A4'].font = Font(name='微软雅黑', size=12, bold=True, color='1F4E78')
                        ws_cover.merge_cells('A4:F4')
                        
                        # KPI表头
                        kpi_headers = ['指标', '数值', '目标', '状态', '指标', '数值']
                        for col, header in enumerate(kpi_headers, 1):
                            cell = ws_cover.cell(5, col, header)
                            cell.font = header_font
                            cell.fill = header_fill
                            cell.alignment = center_align
                            cell.border = thin_border
                        
                        # KPI数据
                        kpi_data = [
                            ['CPK', round(stats['cpk'], 3), f"≥{params['cpk_lower']}", quality_status, '合格率', f"{yield_rate:.2f}%"],
                            ['CP', round(stats['cp'], 3), '≥1.33', '合格' if stats['cp'] >= 1.33 else '不足', '不合格率(估算)', f"{total_defect_rate:.4f}%"],
                            ['PPK', round(ppk, 3), f"≥{params['cpk_lower']}", '合格' if ppk >= params['cpk_lower'] else '不足', '西格玛水平', f"{sigma_level}σ"],
                            ['均值', round(stats['mean'], 4), f"{params['mean']}±0.1", '符合' if abs(stats['mean'] - params['mean']) <= 0.1 else '偏离', '样本数', len(numbers)],
                        ]
                        
                        for row_idx, row_data in enumerate(kpi_data, 6):
                            for col_idx, value in enumerate(row_data, 1):
                                cell = ws_cover.cell(row_idx, col_idx, value)
                                cell.font = bold_font if col_idx in [1, 5] else normal_font
                                cell.border = thin_border
                                cell.alignment = center_align
                                
                                # 状态列着色
                                if col_idx == 4:
                                    if value in ['卓越', '优秀', '合格', '符合', '合格']:
                                        cell.fill = green_fill
                                    elif value in ['警告', '不足']:
                                        cell.fill = yellow_fill
                                    elif value in ['不合格', '偏离']:
                                        cell.fill = red_fill
                                
                                if row_idx % 2 == 0 and col_idx != 4:
                                    cell.fill = light_fill
                        
                        # 质量判定区域
                        ws_cover['A11'] = '质量判定与建议'
                        ws_cover['A11'].font = Font(name='微软雅黑', size=12, bold=True, color='1F4E78')
                        ws_cover.merge_cells('A11:F11')
                        
                        quality_judgement = [
                            ['判定项目', '结果', '等级', '建议措施', '', ''],
                            ['CPK等级判定', quality_status, cpk_level, suggestion, '', ''],
                            ['过程能力', '充足' if stats['cpk'] >= 1.33 else '不足', f"CPK={round(stats['cpk'], 2)}", '可接受，继续监控' if stats['cpk'] >= 1.33 else '需要制定改进计划', '', ''],
                            ['正态性检验', '符合' if normality_result['is_normal'] else '不符合', f"p={normality_result['shapiro_p']:.4f}", '数据可信，可用于分析' if normality_result['is_normal'] else '数据可能非正态，谨慎解读', '', ''],
                        ]
                        
                        for row_idx, row_data in enumerate(quality_judgement, 12):
                            for col_idx, value in enumerate(row_data, 1):
                                cell = ws_cover.cell(row_idx, col_idx, value)
                                cell.font = header_font if row_idx == 12 else normal_font
                                cell.fill = header_fill if row_idx == 12 else (cpk_fill if row_idx == 13 and col_idx == 2 else light_fill if row_idx % 2 == 0 else PatternFill())
                                cell.border = thin_border
                                cell.alignment = center_align if col_idx in [2, 3] else left_align
                                if row_idx == 12:
                                    cell.fill = header_fill
                        
                        # 合并建议措施列
                        for row in [13, 14, 15]:
                            ws_cover.merge_cells(f'D{row}:F{row}')
                        
                        # 报告信息
                        ws_cover['A17'] = '报告信息'
                        ws_cover['A17'].font = Font(name='微软雅黑', size=12, bold=True, color='1F4E78')
                        ws_cover.merge_cells('A17:F17')
                        
                        report_details = [
                            ['生成时间', datetime.now().strftime('%Y-%m-%d %H:%M:%S'), '规格上限', params['upper_spec'], '规格下限', params['lower_spec']],
                            ['软件版本', 'NDG v2.0', '目标值', params['center'], '设定均值', params['mean']],
                            ['分析人员', '', '审核人员', '', '批准人员', ''],
                        ]
                        
                        for row_idx, row_data in enumerate(report_details, 18):
                            for col_idx, value in enumerate(row_data, 1):
                                cell = ws_cover.cell(row_idx, col_idx, value)
                                cell.font = bold_font if col_idx % 2 == 1 else normal_font
                                cell.border = thin_border
                                cell.alignment = left_align if col_idx % 2 == 1 else center_align
                                if row_idx % 2 == 0:
                                    cell.fill = light_fill
                        
                        # 设置列宽
                        ws_cover.column_dimensions['A'].width = 18
                        ws_cover.column_dimensions['B'].width = 15
                        ws_cover.column_dimensions['C'].width = 18
                        ws_cover.column_dimensions['D'].width = 15
                        ws_cover.column_dimensions['E'].width = 18
                        ws_cover.column_dimensions['F'].width = 15
                        
                        # ========== 2. 参数设置 ==========
                        ws_params = wb.create_sheet('参数设置')
                        
                        # 表头
                        headers = ['参数名称', '设定值', '单位', '说明']
                        for col, header in enumerate(headers, 1):
                            cell = ws_params.cell(1, col, header)
                            cell.font = header_font
                            cell.fill = header_fill
                            cell.alignment = center_align
                            cell.border = thin_border
                        
                        # 数据
                        params_rows = [
                            ['目标值（中心值）', params['center'], '-', '规格中心目标'],
                            ['规格上限（USL）', params['upper_spec'], '-', 'Upper Specification Limit'],
                            ['规格下限（LSL）', params['lower_spec'], '-', 'Lower Specification Limit'],
                            ['CPK上限', params['cpk_upper'], '-', '优秀等级标准'],
                            ['CPK下限', params['cpk_lower'], '-', '合格等级标准'],
                            ['设定均值', params['mean'], '-', '正态分布均值'],
                            ['精度（小数位）', params['precision'], '位', '数值精度'],
                            ['样本数量', params['count'], '个', '生成随机数个数'],
                        ]
                        
                        for row_idx, row_data in enumerate(params_rows, 2):
                            for col_idx, value in enumerate(row_data, 1):
                                cell = ws_params.cell(row_idx, col_idx, value)
                                cell.font = normal_font
                                cell.border = thin_border
                                cell.alignment = center_align if col_idx in [2, 3] else left_align
                                if row_idx % 2 == 0:
                                    cell.fill = light_fill
                        
                        # 设置列宽
                        ws_params.column_dimensions['A'].width = 22
                        ws_params.column_dimensions['B'].width = 15
                        ws_params.column_dimensions['C'].width = 10
                        ws_params.column_dimensions['D'].width = 30
                        
                        # ========== 3. 统计结果 ==========
                        ws_stats = wb.create_sheet('统计结果')
                        
                        # 表头
                        headers = ['统计量', '计算值', '标准/目标', '判定结果']
                        for col, header in enumerate(headers, 1):
                            cell = ws_stats.cell(1, col, header)
                            cell.font = header_font
                            cell.fill = header_fill
                            cell.alignment = center_align
                            cell.border = thin_border
                        
                        # 数据
                        target_mean_min = params['mean'] - 0.1
                        target_mean_max = params['mean'] + 0.1
                        mean_pass = target_mean_min <= stats['mean'] <= target_mean_max
                        
                        stats_rows = [
                            ['样本数量（n）', len(numbers), params['count'], '✓ 符合' if len(numbers) == params['count'] else '✗ 不符'],
                            ['均值（Mean）', round(stats['mean'], 4), f"{params['mean']}±0.1", '✓ 符合' if mean_pass else '✗ 偏离'],
                            ['标准差（Std）', round(stats['std'], 4), '-', '-'],
                            ['方差（Var）', round(variance, 4), '-', '-'],
                            ['最大值（Max）', round(max_val, 4), f"≤{params['upper_spec']}", '✓ 合格' if max_val <= params['upper_spec'] else '✗ 超规'],
                            ['最小值（Min）', round(min_val, 4), f"≥{params['lower_spec']}", '✓ 合格' if min_val >= params['lower_spec'] else '✗ 超规'],
                            ['范围（Range）', round(range_val, 4), '-', '-'],
                            ['过程能力指数（CP）', round(stats['cp'], 4), '≥1.33', '✓ 充足' if stats['cp'] >= 1.33 else '✗ 不足'],
                            ['过程能力指数（CPK）', round(stats['cpk'], 4), f"{params['cpk_lower']}-{params['cpk_upper']}", '✓ ' + quality_status],
                        ]
                        
                        for row_idx, row_data in enumerate(stats_rows, 2):
                            for col_idx, value in enumerate(row_data, 1):
                                cell = ws_stats.cell(row_idx, col_idx, value)
                                cell.font = normal_font
                                cell.border = thin_border
                                cell.alignment = center_align
                                
                                # 判定结果列着色
                                if col_idx == 4:
                                    if '✓' in str(value):
                                        cell.fill = green_fill
                                        cell.font = Font(name='微软雅黑', size=10, bold=True, color='006100')
                                    elif '✗' in str(value):
                                        cell.fill = red_fill
                                        cell.font = Font(name='微软雅黑', size=10, bold=True, color='9C0006')
                                
                                if row_idx % 2 == 0 and col_idx != 4:
                                    cell.fill = light_fill
                        
                        # 设置列宽
                        ws_stats.column_dimensions['A'].width = 25
                        ws_stats.column_dimensions['B'].width = 18
                        ws_stats.column_dimensions['C'].width = 20
                        ws_stats.column_dimensions['D'].width = 15
                        
                        # ========== 4. 质量判定 ==========
                        ws_quality = wb.create_sheet('质量判定')
                        
                        # 表头
                        headers = ['判定项目', '结果', '等级/数值', '建议措施']
                        for col, header in enumerate(headers, 1):
                            cell = ws_quality.cell(1, col, header)
                            cell.font = header_font
                            cell.fill = header_fill
                            cell.alignment = center_align
                            cell.border = thin_border
                        
                        # 数据
                        quality_rows = [
                            ['CPK等级判定', quality_status, cpk_level, suggestion],
                            ['过程能力评估', '充足' if stats['cpk'] >= 1.33 else '不足', f"CPK={round(stats['cpk'], 2)}", '可接受' if stats['cpk'] >= 1.33 else '需改进'],
                            ['正态性检验', '符合' if normality_result['is_normal'] else '不符合', f"p={normality_result['shapiro_p']:.4f}", '数据可信' if normality_result['is_normal'] else '谨慎使用'],
                            ['规格符合性', '合格' if (min_val >= params['lower_spec'] and max_val <= params['upper_spec']) else '不合格', '100%' if (min_val >= params['lower_spec'] and max_val <= params['upper_spec']) else '<100%', '全部在规格内' if (min_val >= params['lower_spec'] and max_val <= params['upper_spec']) else '存在超规'],
                        ]
                        
                        for row_idx, row_data in enumerate(quality_rows, 2):
                            for col_idx, value in enumerate(row_data, 1):
                                cell = ws_quality.cell(row_idx, col_idx, value)
                                cell.font = normal_font
                                cell.border = thin_border
                                cell.alignment = center_align if col_idx in [2, 3] else left_align
                                
                                # 结果列着色
                                if col_idx == 2:
                                    if value in ['优秀', '卓越', '合格', '充足', '符合', '合格']:
                                        cell.fill = green_fill
                                        cell.font = Font(name='微软雅黑', size=10, bold=True, color='006100')
                                    elif value in ['警告']:
                                        cell.fill = yellow_fill
                                        cell.font = Font(name='微软雅黑', size=10, bold=True, color='9C5700')
                                    elif value in ['不合格', '不足', '不符合']:
                                        cell.fill = red_fill
                                        cell.font = Font(name='微软雅黑', size=10, bold=True, color='9C0006')
                                
                                if row_idx % 2 == 0:
                                    cell.fill = light_fill
                        
                        # 设置列宽
                        ws_quality.column_dimensions['A'].width = 20
                        ws_quality.column_dimensions['B'].width = 15
                        ws_quality.column_dimensions['C'].width = 20
                        ws_quality.column_dimensions['D'].width = 25
                        
                        # ========== 5. 正态性检验 ==========
                        ws_norm = wb.create_sheet('正态性检验')
                        
                        # 表头
                        headers = ['检验项目', '统计量', 'p值/参考值', '判定标准', '结果']
                        for col, header in enumerate(headers, 1):
                            cell = ws_norm.cell(1, col, header)
                            cell.font = header_font
                            cell.fill = header_fill
                            cell.alignment = center_align
                            cell.border = thin_border
                        
                        # 数据
                        normality_rows = [
                            ['Shapiro-Wilk检验', round(normality_result['shapiro_stat'], 4), round(normality_result['shapiro_p'], 4), 'p > 0.05', '符合' if normality_result['is_normal'] else '不符合'],
                            ['偏度（Skewness）', round(normality_result['skewness'], 4), '-', '-0.5 ~ 0.5', '对称' if abs(normality_result['skewness']) <= 0.5 else '偏斜'],
                            ['峰度（Kurtosis）', round(normality_result['kurtosis'], 4), '-', '-1 ~ 1', '正常' if abs(normality_result['kurtosis']) <= 1 else '异常'],
                            ['Anderson-Darling检验', round(normality_result.get('anderson_stat', 0), 4) if 'anderson_stat' in normality_result else 'N/A', 
                             f"临界值: {normality_result.get('anderson_critical', 'N/A')}", '统计量 < 临界值', '符合' if normality_result.get('anderson_stat', 0) < normality_result.get('anderson_critical', 999) else '不符合'] if 'anderson_stat' in normality_result else ['Anderson-Darling检验', '未计算', '-', '-', '-'],
                        ]
                        
                        for row_idx, row_data in enumerate(normality_rows, 2):
                            for col_idx, value in enumerate(row_data, 1):
                                cell = ws_norm.cell(row_idx, col_idx, value)
                                cell.font = normal_font
                                cell.border = thin_border
                                cell.alignment = center_align
                                
                                # 结果列着色
                                if col_idx == 5:
                                    if value in ['符合', '对称', '正常']:
                                        cell.fill = green_fill
                                        cell.font = Font(name='微软雅黑', size=10, bold=True, color='006100')
                                    elif value in ['不符合', '偏斜', '异常']:
                                        cell.fill = red_fill
                                        cell.font = Font(name='微软雅黑', size=10, bold=True, color='9C0006')
                                
                                if row_idx % 2 == 0 and col_idx != 5:
                                    cell.fill = light_fill
                        
                        # 添加说明
                        ws_norm['A7'] = '说明：'
                        ws_norm['A7'].font = bold_font
                        ws_norm['A8'] = '• Shapiro-Wilk检验：适用于小样本（n < 50）的正态性检验'
                        ws_norm['A9'] = '• 偏度：衡量分布对称性，0为完全对称'
                        ws_norm['A10'] = '• 峰度：衡量分布尖锐程度，0为正态分布'
                        ws_norm['A11'] = '• 显著性水平α = 0.05'
                        for row in range(8, 12):
                            ws_norm[f'A{row}'].font = Font(name='微软雅黑', size=9, italic=True)
                            ws_norm.merge_cells(f'A{row}:E{row}')
                        
                        # 设置列宽
                        ws_norm.column_dimensions['A'].width = 25
                        ws_norm.column_dimensions['B'].width = 15
                        ws_norm.column_dimensions['C'].width = 18
                        ws_norm.column_dimensions['D'].width = 20
                        ws_norm.column_dimensions['E'].width = 15
                        
                        # ========== 6. 原始数据 ==========
                        ws_data = wb.create_sheet('原始数据')
                        
                        # 表头
                        headers = ['样本ID', '测量值', '与目标偏差', '规格判定', '备注']
                        for col, header in enumerate(headers, 1):
                            cell = ws_data.cell(1, col, header)
                            cell.font = header_font
                            cell.fill = header_fill
                            cell.alignment = center_align
                            cell.border = thin_border
                        
                        # 数据
                        for idx, num in enumerate(numbers, 1):
                            deviation = num - params['mean']
                            if num > params['upper_spec']:
                                status = '超上限'
                                note = '超出规格上限'
                            elif num < params['lower_spec']:
                                status = '超下限'
                                note = '低于规格下限'
                            else:
                                status = '合格'
                                note = '-'
                            
                            row_data = [f'S{idx:03d}', round(num, params['precision']), round(deviation, params['precision']), status, note]
                            
                            for col_idx, value in enumerate(row_data, 1):
                                cell = ws_data.cell(idx + 1, col_idx, value)
                                cell.font = normal_font
                                cell.border = thin_border
                                cell.alignment = center_align
                                
                                # 判定列着色
                                if col_idx == 4:
                                    if status == '合格':
                                        cell.fill = green_fill
                                        cell.font = Font(name='微软雅黑', size=10, color='006100')
                                    else:
                                        cell.fill = red_fill
                                        cell.font = Font(name='微软雅黑', size=10, color='9C0006')
                                
                                if (idx + 1) % 2 == 0 and col_idx != 4:
                                    cell.fill = light_fill
                        
                        # 设置列宽
                        ws_data.column_dimensions['A'].width = 12
                        ws_data.column_dimensions['B'].width = 15
                        ws_data.column_dimensions['C'].width = 15
                        ws_data.column_dimensions['D'].width = 12
                        ws_data.column_dimensions['E'].width = 20
                        
                        # ========== 7. 分布分析（带图表） ==========
                        ws_chart = wb.create_sheet('分布分析')
                        
                        # 创建直方图数据
                        hist, bin_edges = np.histogram(numbers, bins=20)
                        bin_centers = (bin_edges[:-1] + bin_edges[1:]) / 2
                        
                        # 写入直方图数据
                        ws_chart['A1'] = '分布分析数据'
                        ws_chart['A1'].font = Font(name='微软雅黑', size=12, bold=True, color='1F4E78')
                        ws_chart.merge_cells('A1:D1')
                        
                        # 直方图数据表头
                        chart_headers = ['组中值', '频数', '频率(%)', '累积频率(%)']
                        for col, header in enumerate(chart_headers, 1):
                            cell = ws_chart.cell(2, col, header)
                            cell.font = header_font
                            cell.fill = header_fill
                            cell.alignment = center_align
                            cell.border = thin_border
                        
                        # 计算累积频率
                        cumulative_freq = 0
                        total_count = len(numbers)
                        
                        for idx, (center, count) in enumerate(zip(bin_centers, hist), 3):
                            freq_percent = (count / total_count) * 100
                            cumulative_freq += freq_percent
                            
                            ws_chart.cell(idx, 1, round(center, 4)).font = normal_font
                            ws_chart.cell(idx, 2, count).font = normal_font
                            ws_chart.cell(idx, 3, round(freq_percent, 2)).font = normal_font
                            ws_chart.cell(idx, 4, round(cumulative_freq, 2)).font = normal_font
                            
                            for col in range(1, 5):
                                ws_chart.cell(idx, col).border = thin_border
                                ws_chart.cell(idx, col).alignment = center_align
                                if idx % 2 == 0:
                                    ws_chart.cell(idx, col).fill = light_fill
                        
                        # 添加统计参考线数据
                        ref_row = len(hist) + 5
                        ws_chart[f'A{ref_row}'] = '统计参考线'
                        ws_chart[f'A{ref_row}'].font = Font(name='微软雅黑', size=11, bold=True, color='1F4E78')
                        ws_chart.merge_cells(f'A{ref_row}:D{ref_row}')
                        
                        ref_data = [
                            ['参考项', '数值', '说明', ''],
                            ['目标值(中心)', params['center'], '规格中心', ''],
                            ['规格上限(USL)', params['upper_spec'], 'Upper Spec Limit', ''],
                            ['规格下限(LSL)', params['lower_spec'], 'Lower Spec Limit', ''],
                            ['实际均值', round(stats['mean'], 4), '样本均值', ''],
                            ['实际标准差', round(stats['std'], 4), '样本标准差', ''],
                        ]
                        
                        for row_idx, row_data in enumerate(ref_data, ref_row + 1):
                            for col_idx, value in enumerate(row_data, 1):
                                cell = ws_chart.cell(row_idx, col_idx, value)
                                cell.font = header_font if row_idx == ref_row + 1 else normal_font
                                cell.fill = header_fill if row_idx == ref_row + 1 else light_fill
                                cell.border = thin_border
                                cell.alignment = center_align if col_idx in [2, 3] else left_align
                        
                        # 创建柱状图
                        chart = BarChart()
                        chart.type = "col"
                        chart.style = 10
                        chart.title = f"{group['name']} - 数据分布直方图"
                        chart.y_axis.title = '频数'
                        chart.x_axis.title = '数值'
                        chart.height = 10
                        chart.width = 18
                        
                        data = Reference(ws_chart, min_col=2, min_row=2, max_row=len(hist) + 2)
                        cats = Reference(ws_chart, min_col=1, min_row=3, max_row=len(hist) + 2)
                        chart.add_data(data, titles_from_data=True)
                        chart.set_categories(cats)
                        
                        ws_chart.add_chart(chart, "F2")
                        
                        # ========== 正态分布曲线数据 ==========
                        # 在直方图数据下方添加正态分布曲线数据
                        curve_start_row = len(hist) + 15
                        
                        ws_chart[f'A{curve_start_row}'] = '正态分布曲线数据'
                        ws_chart[f'A{curve_start_row}'].font = Font(name='微软雅黑', size=12, bold=True, color='1F4E78')
                        ws_chart.merge_cells(f'A{curve_start_row}:C{curve_start_row}')
                        
                        # 正态分布曲线数据表头
                        curve_headers = ['X值', '理论概率密度', '理论频数']
                        for col, header in enumerate(curve_headers, 1):
                            cell = ws_chart.cell(curve_start_row + 1, col, header)
                            cell.font = header_font
                            cell.fill = header_fill
                            cell.alignment = center_align
                            cell.border = thin_border
                        
                        # 生成正态分布曲线数据点
                        x_min = min(numbers) - 0.5
                        x_max = max(numbers) + 0.5
                        x_values = np.linspace(x_min, x_max, 100)
                        bin_width = (x_max - x_min) / 20
                        
                        for idx, x in enumerate(x_values, curve_start_row + 2):
                            # 计算理论概率密度
                            pdf = (1 / (stats['std'] * np.sqrt(2 * np.pi))) * np.exp(-0.5 * ((x - stats['mean']) / stats['std']) ** 2)
                            # 计算理论频数
                            expected_freq = pdf * len(numbers) * bin_width
                            
                            ws_chart.cell(idx, 1, round(x, 4)).font = normal_font
                            ws_chart.cell(idx, 2, round(pdf, 6)).font = normal_font
                            ws_chart.cell(idx, 3, round(expected_freq, 2)).font = normal_font
                            
                            for col in range(1, 4):
                                ws_chart.cell(idx, col).border = thin_border
                                ws_chart.cell(idx, col).alignment = center_align
                                if idx % 2 == 0:
                                    ws_chart.cell(idx, col).fill = light_fill
                        
                        # 创建正态分布曲线图（折线图）
                        line_chart = LineChart()
                        line_chart.style = 12
                        line_chart.title = f"{group['name']} - 正态分布曲线"
                        line_chart.y_axis.title = '频数'
                        line_chart.x_axis.title = '数值'
                        line_chart.height = 10
                        line_chart.width = 18
                        
                        # 使用理论频数数据
                        line_data = Reference(ws_chart, min_col=3, min_row=curve_start_row + 1, max_row=curve_start_row + 100)
                        line_cats = Reference(ws_chart, min_col=1, min_row=curve_start_row + 2, max_row=curve_start_row + 100)
                        line_chart.add_data(line_data, titles_from_data=True)
                        line_chart.set_categories(line_cats)
                        
                        # 设置线条样式（简化版本）
                        if len(line_chart.series) > 0:
                            line_chart.series[0].smooth = True
                        
                        ws_chart.add_chart(line_chart, "F25")
                        
                        # 添加规格线标注
                        spec_row = curve_start_row + 105
                        ws_chart[f'A{spec_row}'] = '规格线位置'
                        ws_chart[f'A{spec_row}'].font = Font(name='微软雅黑', size=11, bold=True, color='1F4E78')
                        ws_chart.merge_cells(f'A{spec_row}:C{spec_row}')
                        
                        spec_data = [
                            ['规格线', 'X值', '说明'],
                            ['规格下限(LSL)', params['lower_spec'], 'Lower Specification Limit'],
                            ['目标值(Target)', params['center'], '规格中心'],
                            ['规格上限(USL)', params['upper_spec'], 'Upper Specification Limit'],
                            ['实际均值', round(stats['mean'], 4), '样本均值'],
                        ]
                        
                        for row_idx, row_data in enumerate(spec_data, spec_row + 1):
                            for col_idx, value in enumerate(row_data, 1):
                                cell = ws_chart.cell(row_idx, col_idx, value)
                                cell.font = header_font if row_idx == spec_row + 1 else normal_font
                                cell.fill = header_fill if row_idx == spec_row + 1 else light_fill
                                cell.border = thin_border
                                cell.alignment = center_align if col_idx in [2, 3] else left_align
                        
                        # 设置列宽
                        ws_chart.column_dimensions['A'].width = 15
                        ws_chart.column_dimensions['B'].width = 12
                        ws_chart.column_dimensions['C'].width = 12
                        ws_chart.column_dimensions['D'].width = 15
                        ws_chart.column_dimensions['E'].width = 3
                        
                        # ========== 8. 完整数据汇总表（带条件参数） ==========
                        ws_full = wb.create_sheet('完整数据汇总')
                        
                        # 标题
                        ws_full['A1'] = f'{group["name"]} - 完整数据汇总表'
                        ws_full['A1'].font = Font(name='微软雅黑', size=14, bold=True, color='1F4E78')
                        ws_full.merge_cells('A1:L1')
                        ws_full['A1'].alignment = center_align
                        
                        # 条件参数表头（第2行）
                        condition_headers = ['目标值', '规格上限', '规格下限', 'CPK上限', 'CPK下限', '设定均值', '精度', '样本数', '实际CPK', '实际CP', '合格率%', '质量等级']
                        for col, header in enumerate(condition_headers, 1):
                            cell = ws_full.cell(2, col, header)
                            cell.font = header_font
                            cell.fill = subheader_fill
                            cell.alignment = center_align
                            cell.border = thin_border
                        
                        # 条件参数值（第3行）
                        condition_values = [
                            params['center'], params['upper_spec'], params['lower_spec'],
                            params['cpk_upper'], params['cpk_lower'], params['mean'],
                            params['precision'], len(numbers),
                            round(stats['cpk'], 3), round(stats['cp'], 3),
                            round(yield_rate, 2), quality_status
                        ]
                        
                        for col, value in enumerate(condition_values, 1):
                            cell = ws_full.cell(3, col, value)
                            cell.font = bold_font
                            cell.alignment = center_align
                            cell.border = thin_border
                            cell.fill = light_fill
                            if col == 12:  # 质量等级着色
                                cell.fill = cpk_fill
                        
                        # 空行
                        # 数据表头（第5行）
                        data_headers = ['序号', '样本ID', '测量值', '与目标偏差', '与均值偏差', 'Z分数', '规格判定', '备注']
                        for col, header in enumerate(data_headers, 1):
                            cell = ws_full.cell(5, col, header)
                            cell.font = header_font
                            cell.fill = header_fill
                            cell.alignment = center_align
                            cell.border = thin_border
                        
                        # 写入数据
                        for idx, num in enumerate(numbers, 1):
                            row = idx + 5
                            deviation_target = num - params['mean']
                            deviation_mean = num - stats['mean']
                            z_score = deviation_mean / stats['std'] if stats['std'] > 0 else 0
                            
                            if num > params['upper_spec']:
                                status = '超上限'
                                note = '超出USL'
                            elif num < params['lower_spec']:
                                status = '超下限'
                                note = '低于LSL'
                            else:
                                status = '合格'
                                note = '-'
                            
                            row_data = [
                                idx, f'S{idx:03d}', round(num, params['precision']),
                                round(deviation_target, params['precision']),
                                round(deviation_mean, params['precision']),
                                round(z_score, 3), status, note
                            ]
                            
                            for col_idx, value in enumerate(row_data, 1):
                                cell = ws_full.cell(row, col_idx, value)
                                cell.font = normal_font
                                cell.border = thin_border
                                cell.alignment = center_align
                                
                                # 判定列着色
                                if col_idx == 7:
                                    if status == '合格':
                                        cell.fill = green_fill
                                        cell.font = Font(name='微软雅黑', size=10, color='006100')
                                    else:
                                        cell.fill = red_fill
                                        cell.font = Font(name='微软雅黑', size=10, color='9C0006')
                                
                                if row % 2 == 0 and col_idx != 7:
                                    cell.fill = light_fill
                        
                        # 设置列宽
                        ws_full.column_dimensions['A'].width = 8
                        ws_full.column_dimensions['B'].width = 12
                        ws_full.column_dimensions['C'].width = 12
                        ws_full.column_dimensions['D'].width = 14
                        ws_full.column_dimensions['E'].width = 14
                        ws_full.column_dimensions['F'].width = 10
                        ws_full.column_dimensions['G'].width = 12
                        ws_full.column_dimensions['H'].width = 15
                        
                        # 保存
                        wb.save(output)
                        output.seek(0)
                        
                        # 使用st.download_button
                        st.download_button(
                            label="下载 Excel",
                            data=output,
                            file_name=f"{group['name']}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                            use_container_width=True
                        )
                        st.success(f"✓ 质量报告已生成：包含8个工作表（执行摘要、参数设置、统计结果、质量判定、正态性检验、原始数据、分布分析[含直方图+正态分布曲线]、完整数据汇总）")
                    
                    # 复制数据
                    if st.button("📋 复制数据", use_container_width=True):
                        numbers_str = ", ".join([f"{num:.{params['precision']}f}" for num in numbers])
                        st.code(numbers_str, language="text")
                
                st.markdown('</div>', unsafe_allow_html=True)
            else:
                st.info("ℹ️ 该条件组尚未生成随机数，请点击上方「生成随机数」按钮")

# ==================== 页脚 ====================
st.markdown("""
<div style="text-align: center; padding: 2rem; color: #95a5a6; font-size: 0.85rem;">
    <hr style="margin-bottom: 1rem; border-color: #ecf0f1;">
    <p>© 2026 NDG - Normal Distribution Generator | 专业质量工程统计分析工具</p>
    <p style="font-size: 0.75rem; margin-top: 0.5rem;">Designed for Quality Engineers</p>
</div>
""", unsafe_allow_html=True)
